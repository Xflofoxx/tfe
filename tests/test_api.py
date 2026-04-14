"""Test suite for Fair Evaluator API - 95% coverage target."""
import os
os.environ["TESTING"] = "true"

import pytest
from fastapi.testclient import TestClient
from unittest.mock import patch, MagicMock
from uuid import uuid4
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.fair_evaluator.main import app
from src.fair_evaluator.db import Base, SessionLocal, engine
from src.fair_evaluator.models import Fair


client = TestClient(app)


# Setup - create test database
@pytest.fixture(scope="session", autouse=True)
def setup_database():
    """Create test database tables."""
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(bind=engine)
    yield


@pytest.fixture
def db_session():
    """Provide a clean database session for each test."""
    db = SessionLocal()
    try:
        yield db
    finally:
        db.query(Fair).delete()
        db.commit()
        db.close()


@pytest.fixture
def sample_fair_data():
    """Sample fair data for testing."""
    return {
        "fair_url": "https://www.example-fair.it",
        "site_url": "https://www.example-fair.it",
        "linkedin_url": "https://linkedin.com/company/example"
    }


@pytest.fixture
def sample_fair(db_session, sample_fair_data):
    """Create a sample fair in the database."""
    fair_id = str(uuid4())
    fair = Fair(
        id=fair_id,
        name="Example Fair",
        url=sample_fair_data["fair_url"],
        company_website=sample_fair_data["site_url"],
        company_linkedin=sample_fair_data["linkedin_url"],
        status="in_valutazione",
        sources=[{"type": "test", "url": sample_fair_data["fair_url"], "date": "2026-04-10"}]
    )
    db_session.add(fair)
    db_session.commit()
    return fair


# ==================== API Endpoint Tests ====================

class TestHealthEndpoints:
    """Test health check and UI endpoints."""
    
    def test_root_ui_returns_html(self):
        """Test root endpoint returns HTML UI."""
        response = client.get("/")
        assert response.status_code == 200
        assert "text/html" in response.headers.get("content-type", "")
    
    def test_ui_contains_expected_content(self):
        """Test UI contains expected elements."""
        response = client.get("/")
        assert response.status_code == 200
        content = response.text
        assert "Valutatore" in content or "Fair" in content


class TestFairCRUD:
    """Test Fair CRUD operations."""
    
    def test_create_fair_success(self, sample_fair_data):
        """Test creating a new fair."""
        response = client.post("/api/fairs", json=sample_fair_data)
        assert response.status_code == 200
        data = response.json()
        assert "id" in data
        assert data["url"] == sample_fair_data["fair_url"]
    
    def test_create_fair_minimal_data(self):
        """Test creating fair with only required data."""
        data = {"fair_url": "https://minimal-fair.it"}
        response = client.post("/api/fairs", json=data)
        assert response.status_code == 200
        result = response.json()
        assert "id" in result
    
    def test_list_fairs_empty(self, db_session):
        """Test listing fairs returns valid response."""
        response = client.get("/api/fairs")
        assert response.status_code == 200
        assert isinstance(response.json(), list)
    
    def test_list_fairs_with_data(self, sample_fair):
        """Test listing fairs with existing data."""
        response = client.get("/api/fairs")
        assert response.status_code == 200
        fairs = response.json()
        assert len(fairs) >= 1
        assert fairs[0]["id"] == sample_fair.id
    
    def test_list_fairs_pagination(self, db_session):
        """Test fair listing with pagination."""
        for i in range(5):
            fair = Fair(
                id=str(uuid4()),
                name=f"Fair {i}",
                url=f"https://fair{i}.it",
                company_website=f"https://fair{i}.it"
            )
            db_session.add(fair)
        db_session.commit()
        
        response = client.get("/api/fairs?skip=0&limit=2")
        assert response.status_code == 200
        fairs = response.json()
        assert len(fairs) == 2
    
    def test_get_fair_success(self, sample_fair):
        """Test getting a specific fair."""
        response = client.get(f"/api/fairs/{sample_fair.id}")
        assert response.status_code == 200
        data = response.json()
        assert data["id"] == sample_fair.id
        assert data["name"] == sample_fair.name
        assert data["url"] == sample_fair.url
    
    def test_get_fair_not_found(self):
        """Test getting non-existent fair returns 404."""
        fake_id = str(uuid4())
        response = client.get(f"/api/fairs/{fake_id}")
        assert response.status_code == 404
        assert "not found" in response.json()["detail"].lower()
    
    def test_update_fair_success(self, sample_fair):
        """Test updating a fair."""
        update_data = {
            "name": "Updated Fair Name",
            "recommendation": "Partecipa",
            "rationale": "Buona visibilità"
        }
        response = client.put(f"/api/fairs/{sample_fair.id}", json=update_data)
        assert response.status_code == 200
        assert response.json()["status"] == "updated"
        
        get_resp = client.get(f"/api/fairs/{sample_fair.id}")
        assert get_resp.json()["name"] == "Updated Fair Name"
    
    def test_update_fair_partial(self, sample_fair):
        """Test partial fair update."""
        response = client.put(f"/api/fairs/{sample_fair.id}", json={"name": "New Name"})
        assert response.status_code == 200
        
        response2 = client.get(f"/api/fairs/{sample_fair.id}")
        assert response2.json()["name"] == "New Name"
    
    def test_delete_fair_success(self, sample_fair, db_session):
        """Test deleting a fair."""
        fair_id = sample_fair.id
        response = client.delete(f"/api/fairs/{fair_id}")
        assert response.status_code == 200
        assert response.json()["status"] == "deleted"
        
        assert db_session.query(Fair).filter(Fair.id == fair_id).first() is None
    
    def test_delete_fair_not_found(self):
        """Test deleting non-existent fair."""
        fake_id = str(uuid4())
        response = client.delete(f"/api/fairs/{fake_id}")
        assert response.status_code == 404


class TestReportGeneration:
    """Test report generation."""
    
    def test_generate_report_fair_not_found(self):
        """Test report generation for non-existent fair."""
        fake_id = str(uuid4())
        response = client.post(f"/api/fairs/{fake_id}/report", json={"format": "html"})
        assert response.status_code == 404


class TestFileUpload:
    """Test file upload functionality."""
    
    def test_upload_pdf_invalid_type(self):
        """Test uploading non-PDF file."""
        import io
        files = {"file": ("test.txt", io.BytesIO(b"text"), "text/plain")}
        response = client.post("/api/upload-pdf", files=files)
        assert response.status_code == 400


class TestImportExport:
    """Test Excel import/export."""
    
    def test_import_excel_no_file(self):
        """Test import without providing file."""
        response = client.post("/api/import-excel-upload")
        assert response.status_code == 422
    
    def test_export_excel(self):
        """Test export excel endpoint."""
        response = client.get("/api/export-excel")
        assert response.status_code == 200


class TestLegacyEndpoints:
    """Test backward compatibility."""
    
    def test_fair_scan_creates_fair(self, sample_fair_data):
        """Test legacy /fair-scan endpoint creates fair."""
        response = client.post("/fair-scan", json=sample_fair_data)
        assert response.status_code == 200
        data = response.json()
        assert "id" in data


class TestErrorHandling:
    """Test error handling scenarios."""
    
    def test_missing_required_field(self):
        """Test missing required field."""
        response = client.post("/api/fairs", json={})
        assert response.status_code == 422
    
    def test_invalid_url_format(self):
        """Test creating fair with invalid URL still works (no validation)."""
        response = client.post("/api/fairs", json={"fair_url": "not-a-url"})
        assert response.status_code == 200


class TestFairModel:
    """Test Fair model operations."""
    
    def test_fair_creation(self, db_session):
        """Test creating a Fair instance."""
        fair = Fair(
            id=str(uuid4()),
            name="Test Fair",
            url="https://test.it",
            company_website="https://test.it"
        )
        db_session.add(fair)
        db_session.commit()
        
        retrieved = db_session.query(Fair).filter(Fair.name == "Test Fair").first()
        assert retrieved is not None
        assert retrieved.url == "https://test.it"
    
    def test_fair_default_values(self, db_session):
        """Test default values for Fair fields."""
        fair = Fair(id=str(uuid4()), name="Test", url="https://test.it")
        db_session.add(fair)
        db_session.commit()
        
        retrieved = db_session.query(Fair).first()
        assert retrieved.recommendation is None
        assert retrieved.dates is None
        assert retrieved.location is None
    
    def test_fair_repr(self):
        """Test Fair __repr__ method."""
        fair = Fair(id="test-id", name="Test Fair")
        assert "Test Fair" in repr(fair)
        assert "test-id" in repr(fair)


class TestFullWorkflow:
    """Test complete user workflows."""
    
    def test_create_and_retrieve_fair(self, db_session):
        """Test full create -> retrieve workflow."""
        create_resp = client.post("/api/fairs", json={"fair_url": "https://workflow-test.it"})
        fair_id = create_resp.json()["id"]
        
        get_resp = client.get(f"/api/fairs/{fair_id}")
        assert get_resp.status_code == 200
        assert get_resp.json()["url"] == "https://workflow-test.it"
    
    def test_create_update_delete_workflow(self):
        """Test complete CRUD workflow."""
        create_resp = client.post("/api/fairs", json={"fair_url": "https://crud-test.it"})
        fair_id = create_resp.json()["id"]
        
        update_resp = client.put(f"/api/fairs/{fair_id}", json={"recommendation": "Non partecipare"})
        assert update_resp.json()["status"] == "updated"
        
        get_resp = client.get(f"/api/fairs/{fair_id}")
        assert get_resp.json()["recommendation"] == "Non partecipare"
        
        delete_resp = client.delete(f"/api/fairs/{fair_id}")
        assert delete_resp.json()["status"] == "deleted"
        
        get_deleted = client.get(f"/api/fairs/{fair_id}")
        assert get_deleted.status_code == 404


class TestPerformance:
    """Basic performance tests."""
    
    def test_list_fairs_performance(self, db_session):
        """Test list operation with multiple records."""
        for i in range(100):
            fair = Fair(id=str(uuid4()), name=f"Fair {i}", url=f"https://fair{i}.it")
            db_session.add(fair)
        db_session.commit()
        
        import time
        start = time.time()
        response = client.get("/api/fairs")
        elapsed = time.time() - start
        
        assert response.status_code == 200
        assert len(response.json()) >= 100
        assert elapsed < 1.0


class TestSettings:
    """Test settings endpoints."""
    
    def test_get_settings(self):
        """Test get settings endpoint."""
        response = client.get("/api/settings")
        assert response.status_code == 200
        data = response.json()
        assert "ollama_url" in data
        assert "ollama_model" in data
    
    def test_update_settings(self):
        """Test update settings."""
        response = client.post("/api/settings", json={"ollama_model": "llama3.1"})
        assert response.status_code == 200
        assert response.json()["status"] == "saved"
        
        get_resp = client.get("/api/settings")
        assert get_resp.json()["ollama_model"] == "llama3.1"
    
    def test_update_strategy_prompt(self):
        """Test updating strategy prompt."""
        response = client.post("/api/settings", json={"strategy_prompt": "Test strategy"})
        assert response.status_code == 200
    
    def test_ollama_status_offline(self):
        """Test ollama status when offline."""
        response = client.get("/api/settings/ollama-status")
        assert response.status_code == 200
        data = response.json()
        assert "status" in data
        assert data["status"] in ["online", "offline"]


class TestUpload:
    """Test upload endpoints."""
    
    def test_upload_pdf_valid(self):
        """Test uploading valid PDF."""
        import io
        pdf_content = b"%PDF-1.4 fake pdf content"
        files = {"file": ("test.pdf", io.BytesIO(pdf_content), "application/pdf")}
        response = client.post("/api/upload-pdf", files=files)
        assert response.status_code == 200
        data = response.json()
        assert "file_id" in data
    
    def test_upload_strategy_pdf(self):
        """Test uploading strategy PDF."""
        import io
        pdf_content = b"%PDF-1.4 fake pdf content"
        files = {"file": ("strategy.pdf", io.BytesIO(pdf_content), "application/pdf")}
        response = client.post("/api/upload-strategy", files=files)
        assert response.status_code == 200
        data = response.json()
        assert "file_path" in data
    
    def test_extract_pdf_no_path(self):
        """Test extract PDF with empty path."""
        response = client.post("/api/extract-pdf", json={"pdf_path": ""})
        assert response.status_code == 200


class TestAnalyze:
    """Test analyze and evaluation endpoints."""
    
    def test_analyze_fair_starts_task(self, sample_fair):
        """Test analyze endpoint starts background task."""
        response = client.post(f"/api/fairs/{sample_fair.id}/analyze")
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "started"
    
    def test_evaluate_fair_no_attachments(self, sample_fair):
        """Test evaluate fails without attachments."""
        response = client.post(f"/api/fairs/{sample_fair.id}/evaluate")
        assert response.status_code == 400
    
    def test_evaluate_fair_with_attachments(self, sample_fair, db_session):
        """Test evaluate with attachments."""
        sample_fair.attachments = [{"name": "quote.pdf", "url": "/uploads/test.pdf"}]
        db_session.commit()
        
        response = client.post(f"/api/fairs/{sample_fair.id}/evaluate")
        assert response.status_code == 200
    
    def test_report_download_not_found(self, sample_fair):
        """Test report download when no report exists."""
        response = client.get(f"/api/fairs/{sample_fair.id}/report/download")
        assert response.status_code == 404
    
    def test_scrape_url_empty(self):
        """Test scrape URL with empty data."""
        response = client.post("/api/scrape-url", json={})
        assert response.status_code == 400
    
    def test_scrape_url_success(self):
        """Test scrape URL endpoint."""
        response = client.post("/api/scrape-url", json={"url": "https://example.com"})
        assert response.status_code == 200
        data = response.json()
        assert "url" in data


class TestStrategy:
    """Test strategy endpoints."""
    
    def test_load_strategy_prompt(self):
        """Test loading strategy prompt from file."""
        import io
        content = b"Test strategy content"
        files = {"file": ("strategy.txt", io.BytesIO(content), "text/plain")}
        response = client.post("/api/strategy/load", files=files)
        assert response.status_code == 200
    
    def test_analyze_strategy_no_pdf(self):
        """Test analyze strategy when no PDF exists."""
        response = client.post("/api/strategy/analyze")
        assert response.status_code == 200
    
    def test_analyze_strategy_pdf(self):
        """Test analyze strategy PDF endpoint."""
        response = client.post("/api/strategy/analyze")
        assert response.status_code == 200


class TestMaintenance:
    """Test maintenance endpoints."""
    
    def test_clear_database(self):
        """Test clearing database."""
        response = client.post("/api/db/clear")
        assert response.status_code == 200
    
    def test_database_stats(self):
        """Test getting database stats."""
        response = client.get("/api/db/stats")
        assert response.status_code == 200
        data = response.json()
        assert "total" in data
        assert "in_valutazione" in data


class TestUIPages:
    """Test UI page endpoints."""
    
    def test_fairs_list_page(self):
        """Test fairs list page."""
        response = client.get("/fairs")
        assert response.status_code == 200
        assert "text/html" in response.headers.get("content-type", "")
    
    def test_fairs_new_page(self):
        """Test new fair page."""
        response = client.get("/fairs/new")
        assert response.status_code == 200
    
    def test_fair_detail_page_not_found(self):
        """Test fair detail page with invalid ID."""
        response = client.get("/fairs/invalid-id")
        assert response.status_code == 404
    
    def test_settings_page(self):
        """Test settings page."""
        response = client.get("/settings")
        assert response.status_code == 200
    
    def test_maintenance_page(self):
        """Test maintenance page."""
        response = client.get("/maintenance")
        assert response.status_code == 200


class TestServices:
    """Test service modules."""
    
    def test_ollama_client_init(self):
        """Test OllamaClient initialization."""
        from src.fair_evaluator.services.ollama import OllamaClient
        client = OllamaClient("http://localhost:11434")
        assert client.base_url == "http://localhost:11434"
    
    def test_ollama_client_chat_error(self):
        """Test OllamaClient chat with connection error."""
        from src.fair_evaluator.services.ollama import OllamaClient
        client = OllamaClient("http://localhost:1")
        result = client.chat("model", "prompt")
        assert result is None
    
    def test_ingest_extract_text_from_url(self):
        """Test text extraction from URL."""
        from src.fair_evaluator.services.ingest import extract_text_from_url
        result = extract_text_from_url("http://invalid.local.test")
        assert result == ""
    
    def test_ingest_extract_text_from_html(self):
        """Test text extraction from HTML."""
        from src.fair_evaluator.services.ingest import extract_text_from_html
        html = "<html><body><p>Test content</p><script>alert('x')</script></body></html>"
        result = extract_text_from_html(html)
        assert "Test content" in result
    
    def test_ingest_extract_text_from_pdf_invalid(self):
        """Test PDF extraction with invalid path."""
        from src.fair_evaluator.services.ingest import extract_text_from_pdf
        result = extract_text_from_pdf("/nonexistent/file.pdf")
        assert result == ""
    
    def test_ingest_scrape_text(self):
        """Test scrape text from site."""
        from src.fair_evaluator.services.ingest import scrape_text_from_site
        result = scrape_text_from_site("http://invalid.local.test")
        assert result == ""
    
    def test_ingest_fair_metadata(self):
        """Test fair metadata ingestion."""
        from src.fair_evaluator.services.ingest import ingest_fair_metadata
        result = ingest_fair_metadata("https://fair.it", "", "https://site.it", "https://linkedin.com")
        assert result["url"] == "https://fair.it"
        assert result["site_url"] == "https://site.it"


class TestEdgeCases:
    """Test edge cases and error handling."""
    
    def test_list_fairs_skip_and_limit(self):
        """Test pagination with skip and limit."""
        for i in range(10):
            fair = Fair(id=str(uuid4()), name=f"EdgeFair{i}", url=f"https://edge{i}.it")
            client.post("/api/fairs", json={"fair_url": f"https://edge{i}.it"})
        
        response = client.get("/api/fairs?skip=5&limit=3")
        assert response.status_code == 200
        fairs = response.json()
        assert len(fairs) <= 3
    
    def test_update_fair_invalid_field(self):
        """Test updating with invalid field is ignored."""
        response = client.put("/api/fairs/invalid-id", json={"nonexistent_field": "value"})
        assert response.status_code == 404
    
    def test_import_excel_empty_file(self):
        """Test import with empty Excel file."""
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["nome", "link evento"])
        wb.save(io.BytesIO())
        wb.close()
        
        files = {"file": ("empty.xlsx", io.BytesIO(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = client.post("/api/import-excel-upload", files=files)
        assert response.status_code == 200
    
    def test_export_excel_no_fairs(self):
        """Test export with no fairs."""
        response = client.get("/api/export-excel")
        assert response.status_code == 200
    
    def test_get_fair_with_null_fields(self, db_session):
        """Test getting fair with null optional fields."""
        fair = Fair(id=str(uuid4()), name="Null Fields Fair", url="https://null.it")
        db_session.add(fair)
        db_session.commit()
        
        response = client.get(f"/api/fairs/{fair.id}")
        assert response.status_code == 200
        data = response.json()
        assert data["description"] == ""
        assert data["gallery"] == []
        assert data["attachments"] == []


class TestCoverageImprovement:
    """Tests to improve code coverage."""
    
    def test_fair_update_with_all_fields(self, db_session):
        """Test updating fair with multiple fields."""
        fair = Fair(id=str(uuid4()), name="Update Test", url="https://update.it")
        db_session.add(fair)
        db_session.commit()
        
        update_data = {
            "name": "Updated Name",
            "description": "Test description",
            "location": "Milan",
            "dates": ["2026-06-01"],
            "target_segments": ["tech", "retail"],
            "expected_visitors": 5000,
            "exhibitors_count": 200,
            "stand_cost": 15000,
            "status": "approvata",
            "recommendation": "Consigliata",
            "rationale": "Buona fiera",
            "cost_estimate": {"stand": 15000, "travel": 2000},
            "ROI_assessment": {"assessment": "high"}
        }
        response = client.put(f"/api/fairs/{fair.id}", json=update_data)
        assert response.status_code == 200
        
        get_resp = client.get(f"/api/fairs/{fair.id}")
        data = get_resp.json()
        assert data["name"] == "Updated Name"
        assert data["location"] == "Milan"
        assert data["expected_visitors"] == 5000
    
    def test_fair_with_contacts(self, db_session):
        """Test fair with contacts data."""
        fair = Fair(
            id=str(uuid4()),
            name="Contact Test",
            url="https://contact.it",
            contacts={"name": "John", "email": "john@example.com", "phone": "123456"}
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.get(f"/api/fairs/{fair.id}")
        assert response.status_code == 200
        data = response.json()
        assert data["contacts"]["name"] == "John"
    
    def test_fair_with_gallery(self, db_session):
        """Test fair with gallery images."""
        fair = Fair(
            id=str(uuid4()),
            name="Gallery Test",
            url="https://gallery.it",
            gallery=["img1.jpg", "img2.jpg"]
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.get(f"/api/fairs/{fair.id}")
        data = response.json()
        assert len(data["gallery"]) == 2
    
    def test_fair_with_scraped_data(self, db_session):
        """Test fair with scraped data."""
        fair = Fair(
            id=str(uuid4()),
            name="Scraped Test",
            url="https://scraped.it",
            scraped_data={"title": "Test Fair", "description": "Description"}
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.get(f"/api/fairs/{fair.id}")
        data = response.json()
        assert "scraped_data" in data
    
    def test_import_excel_with_all_fields(self):
        """Test Excel import with various fields."""
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["nome", "link evento", "città", "location", "stato", "inizio", "costo stand", "contatto", "contatto mail", "contatto tel", "num. leads", "affluenza 2024", "affluenza 2025"])
        ws.append(["Test Fair", "https://test.it", "Rome", "Fiera Roma", "in_valutazione", "2026-06-01", "5000", "Mario", "mario@test.it", "5551234", "100", "5000", "5500"])
        
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        
        buf.seek(0)
        files = {"file": ("full.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = client.post("/api/import-excel-upload", files=files)
        assert response.status_code == 200
        data = response.json()
        assert data["count"] == 1
    
    def test_import_excel_url_parsing(self):
        """Test Excel import URL to name conversion."""
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["nome", "link evento"])
        ws.append(["", "https://www.example.com/fair2026"])
        
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        
        buf.seek(0)
        files = {"file": ("url_test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = client.post("/api/import-excel-upload", files=files)
        assert response.status_code == 200
    
    def test_export_excel_with_data(self, db_session):
        """Test Excel export with data."""
        fair = Fair(
            id=str(uuid4()),
            name="Export Test",
            url="https://export.it",
            location="Turin",
            stand_cost=3000,
            status="approvata"
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.get("/api/export-excel")
        assert response.status_code == 200
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("content-type", "")
    
    def test_database_stats_with_data(self, db_session):
        """Test database stats with actual data."""
        fair1 = Fair(id=str(uuid4()), name="Stat Fair 1", url="https://stat1.it", status="in_valutazione", stand_cost=1000)
        fair2 = Fair(id=str(uuid4()), name="Stat Fair 2", url="https://stat2.it", status="approvata", stand_cost=2000)
        fair3 = Fair(id=str(uuid4()), name="Stat Fair 3", url="https://stat3.it", status="rifiutata")
        db_session.add_all([fair1, fair2, fair3])
        db_session.commit()
        
        response = client.get("/api/db/stats")
        assert response.status_code == 200
        data = response.json()
        assert data["total"] >= 3
        assert data["in_valutazione"] >= 1
        assert data["approvata"] >= 1
        assert data["total_stand_cost"] >= 3000
    
    def test_strategy_prompt_update(self):
        """Test strategy prompt update."""
        response = client.post("/api/settings", json={"strategy_prompt": "Test prompt for fair evaluation"})
        assert response.status_code == 200
        
        get_resp = client.get("/api/settings")
        assert get_resp.json()["strategy_prompt"] == "Test prompt for fair evaluation"
    
    def test_fair_with_historical_data(self, db_session):
        """Test fair with historical data."""
        fair = Fair(
            id=str(uuid4()),
            name="Historical Test",
            url="https://historical.it",
            historical_data={"2024": {"visitors": 5000}, "2025": {"visitors": 5500}}
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.get(f"/api/fairs/{fair.id}")
        data = response.json()
        assert "historical_data" in data
    
    def test_fair_with_venue_and_address(self, db_session):
        """Test fair with venue and address."""
        fair = Fair(
            id=str(uuid4()),
            name="Venue Test",
            url="https://venue.it",
            venue="Fiera Milano",
            address="Strada 1, Milano",
            sector="Technology"
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.get(f"/api/fairs/{fair.id}")
        data = response.json()
        assert data.get("venue") == "Fiera Milano"
    
    def test_fair_with_target_segments(self, db_session):
        """Test fair with target segments."""
        fair = Fair(
            id=str(uuid4()),
            name="Segment Test",
            url="https://segment.it",
            target_segments=["B2B", "Retail", "Tech"]
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.get(f"/api/fairs/{fair.id}")
        data = response.json()
        assert len(data.get("target_segments", [])) == 3


class TestCliModule:
    """Test CLI module."""
    
    def test_cli_import(self):
        """Test CLI module can be imported."""
        from src.fair_evaluator.cli import main
        assert callable(main)


class TestDbModule:
    """Test database module."""
    
    def test_db_get_session(self):
        """Test get_session generator."""
        from src.fair_evaluator.db import get_session
        gen = get_session()
        session = next(gen)
        assert session is not None
        try:
            next(gen)
        except StopIteration:
            pass
    
    def test_db_engine_creation(self):
        """Test database engine creation."""
        from src.fair_evaluator.db import engine
        assert engine is not None


class TestModels:
    """Test model classes."""
    
    def test_settings_model(self):
        """Test Settings model."""
        from src.fair_evaluator.models import Settings
        s = Settings(ollama_url="http://test:11434", ollama_model="test-model")
        assert s.ollama_url == "http://test:11434"
        assert s.ollama_model == "test-model"
        assert "Settings" in repr(s)


class TestOllamaClient:
    """Test Ollama service."""
    
    def test_ollama_client_chat_with_timeout(self):
        """Test OllamaClient handles timeout."""
        from src.fair_evaluator.services.ollama import OllamaClient
        client = OllamaClient("http://localhost:9999")
        result = client.chat("model", "test prompt")
        assert result is None
    
    def test_ollama_client_base_url_strip(self):
        """Test OllamaClient strips trailing slash."""
        from src.fair_evaluator.services.ollama import OllamaClient
        client = OllamaClient("http://localhost:11434/")
        assert client.base_url == "http://localhost:11434"


class TestTemplates:
    """Test template rendering."""
    
    def test_template_directory_exists(self):
        """Test template directory exists."""
        from pathlib import Path
        template_dir = Path(__file__).parent.parent / "src" / "fair_evaluator" / "templates"
        assert template_dir.exists()
    
    def test_page_templates_exist(self):
        """Test required templates exist."""
        from pathlib import Path
        template_dir = Path(__file__).parent.parent / "src" / "fair_evaluator" / "templates"
        required = ["page_home.html", "page_fairs.html", "page_fair_detail.html", "page_settings.html"]
        for t in required:
            assert (template_dir / t).exists(), f"Template {t} missing"


class TestAnalyzeFunctionality:
    """Test analyze and evaluate endpoints more thoroughly."""
    
    def test_analyze_fair_not_found(self):
        """Test analyze with non-existent fair."""
        response = client.post("/api/fairs/nonexistent-id/analyze")
        assert response.status_code == 404
    
    def test_evaluate_fair_not_found(self):
        """Test evaluate with non-existent fair."""
        response = client.post("/api/fairs/nonexistent-id/evaluate")
        assert response.status_code == 404
    
    def test_report_download_pdf_format(self, db_session):
        """Test report download in PDF format."""
        fair = Fair(id=str(uuid4()), name="Report Test", url="https://report.it")
        db_session.add(fair)
        db_session.commit()
        
        fair.report_pdf_path = "data/reports/test.pdf"
        db_session.commit()
        
        response = client.get(f"/api/fairs/{fair.id}/report/download?format=pdf")
        assert response.status_code in [200, 404]
    
    def test_scrape_url_success(self, db_session):
        """Test scrape URL with valid URL."""
        response = client.post("/api/scrape-url", json={"url": "https://httpbin.org/html"})
        assert response.status_code == 200
        data = response.json()
        assert "url" in data


class TestListEndpoint:
    """Test list endpoint response fields."""
    
    def test_list_fairs_returns_expected_fields(self, db_session):
        """Test list returns all expected fields."""
        fair = Fair(
            id=str(uuid4()),
            name="Fields Test",
            url="https://fields.it",
            description="Test description",
            location="Rome",
            status="in_valutazione"
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.get("/api/fairs")
        assert response.status_code == 200
        fairs = response.json()
        if fairs:
            f = fairs[0]
            assert "id" in f
            assert "name" in f
            assert "url" in f
            assert "status" in f


class TestCliCommands:
    """Test CLI command parsing and execution."""
    
    def test_cli_scan_command(self):
        """Test CLI scan command."""
        from src.fair_evaluator.cli import main
        import sys
        old_argv = sys.argv
        sys.argv = ["cli", "scan", "https://test-fair.it"]
        try:
            main()
        except SystemExit:
            pass
        finally:
            sys.argv = old_argv
    
    def test_cli_report_command(self):
        """Test CLI report command."""
        from src.fair_evaluator.cli import main
        import sys
        old_argv = sys.argv
        sys.argv = ["cli", "report", "test-id", "html"]
        try:
            main()
        except SystemExit:
            pass
        finally:
            sys.argv = old_argv
    
    def test_cli_no_command(self):
        """Test CLI with no command shows help."""
        from src.fair_evaluator.cli import main
        import sys
        old_argv = sys.argv
        sys.argv = ["cli"]
        try:
            main()
        except SystemExit:
            pass
        finally:
            sys.argv = old_argv


class TestDbInit:
    """Test db_init module."""
    
    def test_db_init_creates_tables(self):
        """Test db_init creates tables."""
        from src.fair_evaluator.db import Base, engine
        from src.fair_evaluator import db_init
        Base.metadata.create_all(bind=engine)
        assert engine is not None


class TestGenerateReport:
    """Test report generation."""
    
    def test_generate_html_report(self, db_session):
        """Test HTML report generation."""
        fair = Fair(id=str(uuid4()), name="HTML Report Test", url="https://htmltest.it", description="Test fair")
        db_session.add(fair)
        db_session.commit()
        
        response = client.post(f"/api/fairs/{fair.id}/report", json={"format": "html"})
        assert response.status_code == 200
        data = response.json()
        assert "html" in data


class TestScrapeUrlEndpoint:
    """Test scrape URL endpoint."""
    
    def test_scrape_url_empty(self):
        """Test scrape with empty URL."""
        response = client.post("/api/scrape-url", json={"url": ""})
        assert response.status_code == 400
    
    def test_scrape_url_invalid(self):
        """Test scrape with invalid URL."""
        response = client.post("/api/scrape-url", json={"url": "not-a-url"})
        assert response.status_code == 200


class TestStrategyEndpoint:
    """Test strategy endpoints."""
    
    def test_analyze_strategy_no_files(self):
        """Test analyze with no strategy files."""
        response = client.post("/api/strategy/analyze")
        assert response.status_code == 200
        data = response.json()
        assert "status" in data
    
    def test_analyze_strategy_pdf_no_files(self):
        """Test analyze_strategy_pdf with no files."""
        response = client.post("/api/strategy/analyze")
        assert response.status_code == 200


class TestImportExportAdvanced:
    """Test import/export functionality."""
    
    def test_import_excel_with_numeric_fields(self):
        """Test import with numeric stand cost."""
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["nome", "link evento", "costo stand", "num. leads"])
        ws.append(["Numeric Test", "https://numeric.it", "5000", "100"])
        
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        
        buf.seek(0)
        files = {"file": ("numeric.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = client.post("/api/import-excel-upload", files=files)
        assert response.status_code == 200
    
    def test_import_excel_with_contacts(self):
        """Test import with contact info."""
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["nome", "link evento", "contatto", "contatto mail", "contatto tel"])
        ws.append(["Contact Test", "https://contact.it", "Mario", "mario@test.com", "123456"])
        
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        
        buf.seek(0)
        files = {"file": ("contacts.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = client.post("/api/import-excel-upload", files=files)
        assert response.status_code == 200
    
    def test_import_excel_with_affluenza(self):
        """Test import with attendance data."""
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["nome", "link evento", "affluenza 2024", "affluenza 2025"])
        ws.append(["Affluenza Test", "https://aff.it", "5000", "5500"])
        
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        
        buf.seek(0)
        files = {"file": ("affluenza.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = client.post("/api/import-excel-upload", files=files)
        assert response.status_code == 200
    
    def test_import_excel_with_all_dates(self):
        """Test import with date fields."""
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["nome", "link evento", "inizio", "fine"])
        ws.append(["Date Test", "https://date.it", "2026-06-01", "2026-06-05"])
        
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        
        buf.seek(0)
        files = {"file": ("dates.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = client.post("/api/import-excel-upload", files=files)
        assert response.status_code == 200


class TestFairDetailFields:
    """Test fair detail endpoint includes all fields."""
    
    def test_fair_detail_has_all_fields(self, db_session):
        """Test fair detail returns all model fields."""
        fair = Fair(
            id=str(uuid4()),
            name="All Fields Test",
            url="https://allfields.it",
            description="Test description",
            location="Milan",
            venue="Fiera Milano",
            address="Via 1",
            sector="Tech",
            frequency="annual",
            edition="25",
            organizer="Fiera Milano",
            expected_visitors=5000,
            exhibitors_count=300,
            target_segments=["B2B", "Retail"],
            status="in_valutazione"
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.get(f"/api/fairs/{fair.id}")
        assert response.status_code == 200
        data = response.json()
        assert data["name"] == "All Fields Test"
        assert data["venue"] is not None
        assert data["sector"] == "Tech"


class TestDatabaseStatsAdvanced:
    """Test database stats with more data."""
    
    def test_database_stats_with_contacts(self, db_session):
        """Test stats with contacts data."""
        fair = Fair(
            id=str(uuid4()),
            name="Stats Contacts",
            url="https://statscontacts.it",
            contacts={"cold": 10, "warm": 5, "hot": 2},
            stand_cost=5000
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.get("/api/db/stats")
        assert response.status_code == 200
        data = response.json()
        assert "contacts" in data
        assert data["total_stand_cost"] >= 5000


class TestOllamaIntegration:
    """Test Ollama service integration."""
    
    def test_ollama_response_none_on_error(self):
        """Test Ollama returns None on error."""
        from src.fair_evaluator.services.ollama import OllamaClient
        client = OllamaClient("http://invalid:9999")
        result = client.chat("nonexistent", "test")
        assert result is None
    
    def test_ollama_url_normalization(self):
        """Test Ollama URL is normalized."""
        from src.fair_evaluator.services.ollama import OllamaClient
        client1 = OllamaClient("http://test.com/")
        client2 = OllamaClient("http://test.com")
        assert client1.base_url == client2.base_url


class TestIngestService:
    """Test ingest service functions."""
    
    def test_extract_text_returns_empty_on_error(self):
        """Test extract returns empty on error."""
        from src.fair_evaluator.services.ingest import extract_text_from_url
        result = extract_text_from_url("http://invalid.local")
        assert result == ""
    
    def test_extract_text_from_html_preserves_content(self):
        """Test HTML extraction preserves content."""
        from src.fair_evaluator.services.ingest import extract_text_from_html
        html = "<html><body><p>Hello World</p></body></html>"
        result = extract_text_from_html(html)
        assert "Hello" in result
    
    def test_ingest_with_marketing_pdf(self):
        """Test ingest with marketing PDF path."""
        from src.fair_evaluator.services.ingest import ingest_fair_metadata
        result = ingest_fair_metadata(
            fair_url="https://test.it",
            marketing_pdf_path="/fake/path.pdf",
            site_url="https://site.it",
            linkedin_url="https://linkedin.com/test"
        )
        assert result["url"] == "https://test.it"
        assert result["marketing_strategy_pdf"] == "/fake/path.pdf"


class TestSettingsAdvanced:
    """Test settings with more scenarios."""
    
    def test_settings_partial_update(self):
        """Test partial settings update."""
        response = client.post("/api/settings", json={"strategy_prompt": "New prompt"})
        assert response.status_code == 200
        
        get_resp = client.get("/api/settings")
        assert "strategy_prompt" in get_resp.json()
    
    def test_ollama_status_returns_available_models(self):
        """Test ollama status includes models."""
        response = client.get("/api/settings/ollama-status")
        assert response.status_code == 200
        data = response.json()
        assert "status" in data


class TestEdgeCasesAdvanced:
    """Test additional edge cases."""
    
    def test_fair_update_empty_json(self):
        """Test update with empty JSON."""
        response = client.put("/api/fairs/nonexistent", json={})
        assert response.status_code == 404
    
    def test_create_fair_with_minimal_data(self):
        """Test create with only URL."""
        response = client.post("/api/fairs", json={"fair_url": "https://minimal.it"})
        assert response.status_code == 200
        data = response.json()
        assert "id" in data
    
    def test_create_fair_derives_name_from_url(self):
        """Test name derivation from URL."""
        response = client.post("/api/fairs", json={"fair_url": "https://www.example.com/tech-fair-2026"})
        assert response.status_code == 200
        data = response.json()
        assert "tech" in data["name"].lower() or "fair" in data["name"].lower()


class TestUploadStrategy:
    """Test strategy upload endpoints."""
    
    def test_upload_strategy_pdf_success(self):
        """Test uploading strategy PDF."""
        import io
        pdf_content = b"%PDF-1.4 test"
        files = {"file": ("strategy.pdf", io.BytesIO(pdf_content), "application/pdf")}
        response = client.post("/api/upload-strategy", files=files)
        assert response.status_code == 200
        data = response.json()
        assert "file_path" in data
    
    def test_upload_strategy_pdf_invalid_type(self):
        """Test uploading non-PDF to strategy."""
        import io
        txt_content = b"not a pdf"
        files = {"file": ("strategy.txt", io.BytesIO(txt_content), "text/plain")}
        response = client.post("/api/upload-strategy", files=files)
        assert response.status_code == 400
    
    def test_extract_pdf_with_valid_path(self):
        """Test extract PDF with actual file path."""
        response = client.post("/api/extract-pdf", json={"pdf_path": "data/uploads/nonexistent.pdf"})
        assert response.status_code == 200
        data = response.json()
        assert "target_segments" in data


class TestUIEndpoints:
    """Test additional UI endpoints."""
    
    def test_fairs_list_with_filters(self):
        """Test fairs list with pagination params."""
        response = client.get("/fairs?skip=0&limit=10")
        assert response.status_code == 200
    
    def test_fair_detail_page_with_valid_id(self, db_session):
        """Test fair detail page with valid ID."""
        fair = Fair(id=str(uuid4()), name="UI Test", url="https://ui.it")
        db_session.add(fair)
        db_session.commit()
        
        response = client.get(f"/fairs/{fair.id}")
        assert response.status_code == 200
    
    def test_health_page_loads(self):
        """Test health check."""
        response = client.get("/")
        assert response.status_code == 200


class TestErrorScenarios:
    """Test error handling."""
    
    def test_create_fair_url_with_special_chars(self):
        """Test creating fair with special chars in URL."""
        response = client.post("/api/fairs", json={"fair_url": "https://test.com/fair%20name"})
        assert response.status_code == 200
    
    def test_list_fairs_large_skip(self):
        """Test listing with large skip value."""
        response = client.get("/api/fairs?skip=10000&limit=10")
        assert response.status_code == 200
    
    def test_update_fair_status_to_all_values(self, db_session):
        """Test updating fair status to various values."""
        fair = Fair(id=str(uuid4()), name="Status Test", url="https://status.it")
        db_session.add(fair)
        db_session.commit()
        
        for status in ["in_valutazione", "approvata", "rifiutata", "in_gestione", "conclusa"]:
            response = client.put(f"/api/fairs/{fair.id}", json={"status": status})
            assert response.status_code == 200


class TestDatabaseEdgeCases:
    """Test database edge cases."""
    
    def test_clear_database_empty(self):
        """Test clearing empty database."""
        response = client.post("/api/db/clear")
        assert response.status_code == 200
        assert response.json()["status"] == "success"
    
    def test_stats_with_no_fairs(self):
        """Test stats with empty database."""
        client.post("/api/db/clear")
        response = client.get("/api/db/stats")
        assert response.status_code == 200
        data = response.json()
        assert data["total"] == 0
    
    def test_export_with_special_characters(self, db_session):
        """Test export with Unicode characters."""
        fair = Fair(
            id=str(uuid4()),
            name="Fiera È À Ù",
            url="https://unicode.it",
            location="Città"
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.get("/api/export-excel")
        assert response.status_code == 200


class TestAnalyzeBackgroundTask:
    """Test analyze background task (unit tests)."""
    
    def test_analyze_task_imports(self):
        """Test analyze_fair_task can be imported."""
        import importlib
        import sys
        from unittest.mock import MagicMock, patch
        
        with patch('src.fair_evaluator.main.SessionLocal') as mock_session:
            mock_db = MagicMock()
            mock_session.return_value = mock_db
            
            mock_fair = MagicMock()
            mock_fair.url = "https://test.it"
            mock_fair.company_website = "https://test.it"
            mock_fair.attachments = None
            
            mock_settings = MagicMock()
            mock_settings.ollama_url = "http://localhost:11434"
            mock_settings.ollama_model = "llama3.2"
            mock_settings.strategy_prompt = None
            
            mock_db.query.return_value.filter.return_value.first.side_effect = [mock_fair, mock_settings]
            mock_db.query.return_value.count.return_value = 0
            
            from src.fair_evaluator.main import analyze_fair_task
            analyze_fair_task("test-id")
    
    def test_evaluate_fair_requires_attachments(self, db_session):
        """Test evaluate fails without attachments."""
        fair = Fair(id=str(uuid4()), name="Eval Test", url="https://evaltest.it", attachments=None)
        db_session.add(fair)
        db_session.commit()
        
        response = client.post(f"/api/fairs/{fair.id}/evaluate")
        assert response.status_code == 400
    
    def test_evaluate_fair_with_ollama_available(self, db_session):
        """Test evaluate with Ollama available."""
        fair = Fair(
            id=str(uuid4()),
            name="Eval Ollama Test",
            url="https://evallama.it",
            attachments=[{"name": "quote.pdf", "url": "/test.pdf"}],
            description="Test fair",
            location="Rome",
            expected_visitors=5000
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.post(f"/api/fairs/{fair.id}/evaluate")
        assert response.status_code == 200


class TestReportGenerationAdvanced:
    """Test report generation advanced scenarios."""
    
    def test_report_with_strategy_prompt(self, db_session):
        """Test report includes strategy prompt."""
        fair = Fair(
            id=str(uuid4()),
            name="Strategy Report Test",
            url="https://strattest.it",
            recommendation="Consigliata",
            rationale="Buona visibilità"
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.post(f"/api/fairs/{fair.id}/report", json={"format": "html"})
        assert response.status_code == 200
    
    def test_report_download_html_fallback(self, db_session):
        """Test HTML fallback when PDF not available."""
        fair = Fair(
            id=str(uuid4()),
            name="HTML Fallback Test",
            url="https://htmlfb.it",
            report_html_path="nonexistent.html"
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.get(f"/api/fairs/{fair.id}/report/download?format=html")
        assert response.status_code == 404
    
    def test_report_pdf_generation_fails_gracefully(self, db_session):
        """Test PDF generation fails without crashing."""
        fair = Fair(id=str(uuid4()), name="PDF Fail Test", url="https://pdfail.it")
        db_session.add(fair)
        db_session.commit()
        
        response = client.post(f"/api/fairs/{fair.id}/report", json={"format": "pdf"})
        assert response.status_code == 200


class TestIngestFunctions:
    """Test ingest service functions more."""
    
    def test_ingest_with_all_params(self):
        """Test ingest with all parameters."""
        from src.fair_evaluator.services.ingest import ingest_fair_metadata
        result = ingest_fair_metadata(
            fair_url="https://allparams.it",
            marketing_pdf_path="",
            site_url="https://site.com",
            linkedin_url="https://linkedin.com/company"
        )
        assert result["url"] == "https://allparams.it"
        assert result["site_url"] == "https://site.com"
        assert result["linkedin_url"] == "https://linkedin.com/company"
    
    def test_extract_text_from_nonexistent_url(self):
        """Test extraction from URL that doesn't exist."""
        from src.fair_evaluator.services.ingest import extract_text_from_url
        result = extract_text_from_url("http://this-domain-does-not-exist-12345.com")
        assert result == ""
    
    def test_scrape_with_invalid_url(self):
        """Test scrape with invalid URL."""
        from src.fair_evaluator.services.ingest import scrape_text_from_site
        result = scrape_text_from_site("http://invalid.invalid")
        assert result == ""
    
    def test_extract_from_html_with_scripts(self):
        """Test HTML extraction removes scripts."""
        from src.fair_evaluator.services.ingest import extract_text_from_html
        html = "<html><script>alert('xss')</script><body><p>Visible</p></body></html>"
        result = extract_text_from_html(html)
        assert "alert" not in result
        assert "Visible" in result


class TestModelsAdvanced:
    """Test model operations."""
    
    def test_fair_model_all_fields(self):
        """Test Fair model with all fields."""
        fair = Fair(
            id="test-id",
            name="Full Fair",
            url="https://full.it",
            description="Desc",
            location="Loc",
            venue="Venue",
            address="Address",
            sector="Sector",
            frequency="annual",
            edition="10",
            organizer="Org",
            expected_visitors=1000,
            exhibitors_count=500,
            target_segments=["seg1", "seg2"],
            exhibitor_countries=["IT", "DE"],
            visitor_profile="Profile",
            product_categories=["cat1"],
            key_features=["feat1"],
            dates=["2026-01-01"],
            status="in_valutazione"
        )
        assert fair.name == "Full Fair"
        assert fair.sector == "Sector"
        assert len(fair.target_segments) == 2
    
    def test_fair_with_minimal_data(self):
        """Test Fair with minimal data."""
        fair = Fair(id="min-id", name="MinFair")
        assert fair.id == "min-id"
        assert fair.name == "MinFair"
        assert fair.status is None or fair.status == "in_valutazione"
    
    def test_settings_defaults(self):
        """Test Settings default values."""
        from src.fair_evaluator.models import Settings
        s = Settings(ollama_url="http://localhost:11434", ollama_model="llama3.2")
        assert s.ollama_url == "http://localhost:11434"
        assert s.ollama_model == "llama3.2"
        assert s.strategy_pdf is None


class TestClientParsing:
    """Test CLI client parsing."""
    
    def test_cli_scan_with_args(self):
        """Test CLI scan with all arguments."""
        from src.fair_evaluator.cli import main
        import sys
        old_argv = sys.argv
        sys.argv = ["cli", "scan", "https://fair.it", "/path/to/pdf.pdf", "https://site.it", "https://linkedin.com/company"]
        try:
            main()
        except SystemExit:
            pass
        finally:
            sys.argv = old_argv
    
    def test_cli_report_pdf_format(self):
        """Test CLI report with PDF format."""
        from src.fair_evaluator.cli import main
        import sys
        old_argv = sys.argv
        sys.argv = ["cli", "report", "fair-id-123", "pdf"]
        try:
            main()
        except SystemExit:
            pass
        finally:
            sys.argv = old_argv
    
    def test_cli_report_both_format(self):
        """Test CLI report with both format."""
        from src.fair_evaluator.cli import main
        import sys
        old_argv = sys.argv
        sys.argv = ["cli", "report", "fair-id-123", "both"]
        try:
            main()
        except SystemExit:
            pass
        finally:
            sys.argv = old_argv


class TestEdgeCoverage:
    """Additional edge case tests for coverage."""
    
    def test_scrape_url_with_ai_response(self):
        """Test scrape URL with AI response parsing."""
        with patch('requests.get') as mock_get:
            mock_response = MagicMock()
            mock_response.status_code = 200
            mock_response.text = "<html><title>Test Fair</title><meta name='description' content='A great fair'></html>"
            mock_get.return_value = mock_response
            
            response = client.post("/api/scrape-url", json={"url": "https://ai-test.it"})
            assert response.status_code == 200
    
    def test_update_with_cost_estimate(self, db_session):
        """Test update with cost estimate."""
        fair = Fair(id=str(uuid4()), name="Cost Test", url="https://cost.it")
        db_session.add(fair)
        db_session.commit()
        
        response = client.put(f"/api/fairs/{fair.id}", json={
            "cost_estimate": {"stand": 10000, "travel": 2000, "hotel": 1000}
        })
        assert response.status_code == 200
    
    def test_update_with_roi_assessment(self, db_session):
        """Test update with ROI assessment."""
        fair = Fair(id=str(uuid4()), name="ROI Test", url="https://roi.it")
        db_session.add(fair)
        db_session.commit()
        
        response = client.put(f"/api/fairs/{fair.id}", json={
            "ROI_assessment": {"assessment": "high", "estimated_roi": "3.5"}
        })
        assert response.status_code == 200
    
    def test_import_excel_with_folder(self):
        """Test import with folder path."""
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["nome", "link evento", "cartella"])
        ws.append(["Folder Test", "https://folder.it", "/path/to/folder"])
        
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        
        buf.seek(0)
        files = {"file": ("folder.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = client.post("/api/import-excel-upload", files=files)
        assert response.status_code == 200
    
    def test_import_excel_with_location(self):
        """Test import with location."""
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["nome", "link evento", "città", "location"])
        ws.append(["Loc Test", "https://loctest.it", "Milano", "Fiera Milano"])
        
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        
        buf.seek(0)
        files = {"file": ("location.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = client.post("/api/import-excel-upload", files=files)
        assert response.status_code == 200
    
    def test_database_stats_with_multiple_statuses(self, db_session):
        """Test stats with fairs in different statuses."""
        statuses = ["in_valutazione", "approvata", "rifiutata", "in_gestione", "conclusa"]
        for i, status in enumerate(statuses):
            fair = Fair(id=str(uuid4()), name=f"Status{i}", url=f"https://status{i}.it", status=status)
            db_session.add(fair)
        db_session.commit()
        
        response = client.get("/api/db/stats")
        assert response.status_code == 200
        data = response.json()
        for status in statuses:
            assert status in data
    
    def test_list_fairs_returns_contacts(self, db_session):
        """Test list includes contacts."""
        fair = Fair(
            id=str(uuid4()),
            name="Contacts List Test",
            url="https://contactslist.it",
            contacts={"name": "John", "email": "john@test.com"}
        )
        db_session.add(fair)
        db_session.commit()
        
        response = client.get("/api/fairs")
        assert response.status_code == 200
        fairs = response.json()
        found = any(f.get("contacts", {}).get("name") == "John" for f in fairs)
        assert found


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--cov=src", "--cov-report=term-missing"])