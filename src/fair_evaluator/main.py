import asyncio
import logging
from datetime import datetime
from pathlib import Path
from uuid import uuid4

import requests
import urllib3
from fastapi import BackgroundTasks, Depends, FastAPI, File, HTTPException, UploadFile, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from .db import Base, SessionLocal, engine
from .models import Fair, Settings, Contact, FairAnalysis, OfferComponent, CommercialProposal
from .services.ollama import OllamaClient

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logger = logging.getLogger(__name__)

app = FastAPI(title="Fiera Evaluator (Locale)", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

notifications: list = []
notification_id = 0

def broadcast_notification(message: str, notification_type: str = "info", fair_id: str = None):
    global notification_id
    notification_id += 1
    notif = {
        "id": notification_id,
        "message": message,
        "type": notification_type,
        "fair_id": fair_id,
        "timestamp": datetime.now().isoformat()
    }
    notifications.append(notif)
    if len(notifications) > 100:
        notifications.pop(0)

Base.metadata.create_all(bind=engine)

UPLOAD_DIR = Path("./data/uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
REPORTS_DIR = Path("./data/reports")
REPORTS_DIR.mkdir(parents=True, exist_ok=True)
STRATEGY_DIR = Path("./data/strategy")
STRATEGY_DIR.mkdir(parents=True, exist_ok=True)
PROPOSALS_DIR = Path("./data/proposals")
PROPOSALS_DIR.mkdir(parents=True, exist_ok=True)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def get_settings_db(db: Session = Depends(get_db)):
    settings = db.query(Settings).first()
    if not settings:
        settings = Settings(ollama_url="http://localhost:11434", ollama_model="llama3.2")
        db.add(settings)
        db.commit()
        db.refresh(settings)
    return settings


class FairCreate(BaseModel):
    fair_url: str
    site_url: str | None = ""
    linkedin_url: str | None = ""
    fair_email: str | None = ""
    folder_path: str | None = ""
    network_path: str | None = ""
    name: str | None = None
    year: int | None = None
    location: str | None = None
    dates: list[str] | None = None
    venue: str | None = None
    address: str | None = None
    sector: str | None = None
    organizer: str | None = None
    frequency: str | None = None
    edition: str | None = None
    expected_visitors: int | None = None
    exhibitors_count: int | None = None
    stand_cost: int | None = None
    exhibitor_countries: list[str] | None = None
    visitor_profile: str | None = None
    target_segments: list[str] | None = None
    product_categories: list[str] | None = None
    key_features: list[str] | None = None
    description: str | None = None
    status: str | None = None
    instagram: str | None = None
    facebook: str | None = None
    tiktok: str | None = None
    contacts: dict | None = None
    web_sources: list | None = None
    extraction_regions: list | None = None


class FairUpdate(BaseModel):
    name: str | None = None
    year: int | None = None
    fair_url: str | None = None
    description: str | None = None
    folder_path: str | None = None
    network_path: str | None = None
    dates: list[str] | None = None
    location: str | None = None
    target_segments: list[str] | None = None
    expected_visitors: int | None = None
    exhibitors_count: int | None = None
    sources: list | None = None
    web_sources: list | None = None
    fair_email: str | None = None
    stand_cost: int | None = None
    contacts: dict | None = None
    attachments: list | None = None
    status: str | None = None
    archived: str | None = None
    recommendation: str | None = None
    rationale: str | None = None
    cost_estimate: dict | None = None
    ROI_assessment: dict | None = None
    historical_data: dict | None = None
    venue: str | None = None
    address: str | None = None
    sector: str | None = None
    organizer: str | None = None
    frequency: str | None = None
    edition: str | None = None
    exhibitor_countries: list[str] | None = None
    visitor_profile: str | None = None
    product_categories: list[str] | None = None
    key_features: list[str] | None = None
    instagram: str | None = None
    facebook: str | None = None
    tiktok: str | None = None
    previous_editions: list | None = None


class SettingsUpdate(BaseModel):
    ollama_url: str | None = None
    ollama_model: str | None = None
    strategy_prompt: str | None = None
    default_network_path: str | None = None


@app.get("/api/settings")
def get_settings_endpoint(db: Session = Depends(get_db)):
    s = get_settings_db(db)
    return {"ollama_url": s.ollama_url, "ollama_model": s.ollama_model, "strategy_prompt": s.strategy_prompt or "", "default_network_path": s.default_network_path or ""}


@app.post("/api/settings")
def update_settings(settings: SettingsUpdate, db: Session = Depends(get_db)):
    s = get_settings_db(db)
    if settings.ollama_url is not None:
        s.ollama_url = settings.ollama_url
    if settings.ollama_model is not None:
        s.ollama_model = settings.ollama_model
    if settings.strategy_prompt is not None:
        s.strategy_prompt = settings.strategy_prompt
    if settings.default_network_path is not None:
        s.default_network_path = settings.default_network_path
    db.commit()
    return {"status": "saved"}


@app.get("/api/settings/ollama-status")
def check_ollama_status(db: Session = Depends(get_db)):
    s = get_settings_db(db)
    try:
        resp = requests.get(f"{s.ollama_url}/api/tags", timeout=3)
        if resp.status_code == 200:
            models = resp.json().get("models", [])
            return {"status": "online", "available_models": [m.get("name", "") for m in models], "current_model": s.ollama_model}
    except Exception:
        pass
    return {"status": "offline", "available_models": []}


@app.get("/api/notifications/stream")
def notifications_stream(request):
    async def event_generator():
        last_id = 0
        while True:
            await asyncio.sleep(2)
            notifs = [n for n in notifications if n["id"] > last_id]
            for n in notifs:
                last_id = n["id"]
                yield f"data: {n}\n\n"
    
    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.get("/api/notifications")
def get_notifications():
    return notifications[-20:]


class ScanFolderRequest(BaseModel):
    folder_path: str


@app.post("/api/scan-network-folder")
def scan_network_folder(data: ScanFolderRequest, db: Session = Depends(get_db)):
    import os
    import re
    folder = data.folder_path
    if not folder:
        raise HTTPException(status_code=400, detail="Folder path required")
    
    found_count = 0
    errors = []
    
    try:
        if not os.path.exists(folder):
            try:
                import pathlib
                path = pathlib.Path(folder)
                if not path.exists():
                    return {"found": 0, "errors": ["Percorso non accessibile"]}
            except Exception as e:
                return {"found": 0, "errors": [str(e)]}
        
        fair_name_pattern = re.compile(r"^(.+?)\s*(\d{4})?\s*$", re.IGNORECASE)
        
        for entry in os.listdir(folder):
            entry_path = os.path.join(folder, entry)
            if os.path.isdir(entry_path):
                fair_name = entry.strip()
                year_match = fair_name_pattern.match(fair_name)
                if year_match:
                    fair_name = year_match.group(1).strip()
                    year = int(year_match.group(2)) if year_match.group(2) else datetime.now().year
                else:
                    year = datetime.now().year
                
                existing = db.query(Fair).filter(Fair.name.ilike(fair_name), Fair.year == year).first()
                if existing:
                    errors.append(f"Fiera già esistente: {fair_name} {year}")
                    continue
                
                new_fair = Fair(
                    id=str(uuid4()),
                    name=fair_name,
                    year=year,
                    network_path=entry_path,
                    folder_path=entry_path,
                    status="in_valutazione",
                    sources=[{"type": "network_scan", "date": datetime.now().isoformat()}]
                )
                db.add(new_fair)
                found_count += 1
                
                for subfile in os.listdir(entry_path):
                    subpath = os.path.join(entry_path, subfile)
                    if os.path.isfile(subpath):
                        ext = os.path.splitext(subfile)[1].lower()
                        if ext in ['.pdf', '.xlsx', '.docx', '.txt', '.md']:
                            if not new_fair.attachments:
                                new_fair.attachments = []
                            new_fair.attachments.append({
                                "name": subfile,
                                "url": subpath,
                                "type": ext
                            })
        
        db.commit()
        return {"found": found_count, "errors": errors[:10]}
    
    except Exception as e:
        return {"found": found_count, "errors": [str(e)]}


@app.post("/api/sync-from-network-folder")
def sync_from_network_folder(db: Session = Depends(get_db)):
    """Sincronizza le fiere esistenti con la cartella di rete configurata."""
    settings = db.query(Settings).first()
    if not settings or not settings.default_network_path:
        raise HTTPException(status_code=400, detail="Nessuna cartella configurata")
    
    folder = settings.default_network_path
    return scan_network_folder(ScanFolderRequest(folder_path=folder), db)


@app.post("/api/upload-strategy")
async def upload_strategy(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are allowed")
    file_id = str(uuid4())
    file_path = STRATEGY_DIR / f"{file_id}_{file.filename}"
    with open(file_path, "wb") as f:
        f.write(await file.read())
    s = get_settings_db(db)
    s.strategy_pdf = str(file_path)
    db.commit()
    return {"file_path": str(file_path), "filename": file.filename}


class ExtractPdfRequest(BaseModel):
    pdf_path: str


@app.post("/api/extract-pdf")
def extract_pdf_data(data: ExtractPdfRequest, db: Session = Depends(get_db)):
    import os
    import fitz
    result = {"target_segments": [], "budget_range": None, "marketing_goals": []}
    pdf_path = data.pdf_path
    if not os.path.isabs(pdf_path):
        pdf_path = os.path.join(os.getcwd(), pdf_path)
    pdf_path = os.path.normpath(pdf_path)
    try:
        doc = fitz.open(pdf_path)
        text = "\n".join([page.get_text() for page in doc])
        text_lower = text.lower()

        segments = []
        keywords = ["settore", "target", "segmento", "cliente", "azienda", "mercato", "pubblico"]
        for kw in keywords:
            if kw in text_lower:
                idx = text_lower.find(kw)
                snippet = text[idx:idx+200]
                if '\n' in snippet:
                    snippet = snippet[:snippet.find('\n')]
                if len(snippet) > 10:
                    segments.append(snippet.strip())
        result["target_segments"] = segments[:3]

        budget_keywords = ["budget", "eur", "euro", "costo", "spesa", "investimento"]
        import re
        for kw in budget_keywords:
            if kw in text_lower and "€" in text:
                matches = re.findall(r'\d+\.?\d*\s*(?:€|eur|euro)', text, re.IGNORECASE)
                if matches:
                    result["budget_range"] = f"€ {matches[0]}"
                    break

        goal_keywords = ["obiettivo", "goal", "brand", "lead", "visibilita", "network", "vendita"]
        for kw in goal_keywords:
            if kw in text_lower:
                idx = text_lower.find(kw)
                snippet = text[idx:idx+150]
                if '\n' in snippet:
                    snippet = snippet[:snippet.find('\n')]
                if len(snippet) > 10:
                    result["marketing_goals"].append(snippet.strip())
        result["marketing_goals"] = result["marketing_goals"][:3]
    except Exception as e:
        result["error"] = str(e)
    return result


@app.post("/api/upload-pdf")
async def upload_pdf(file: UploadFile = File(...)):
    if not file.filename or not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are allowed")
    file_id = str(uuid4())
    file_path = UPLOAD_DIR / f"{file_id}_{file.filename}"
    with open(file_path, "wb") as f:
        f.write(await file.read())
    return {"file_path": str(file_path), "filename": file.filename, "file_id": file_id}


@app.post("/api/fairs", response_model=dict)
def create_fair(fair_data: FairCreate, db: Session = Depends(get_db)):
    fair_id = str(uuid4())
    name = fair_data.fair_url
    if '://' in fair_data.fair_url:
        parts = fair_data.fair_url.split('/')
        name = parts[-1] if parts[-1] else parts[-2]
        if '.' in name:
            name = name.split('.')[0]

    year = fair_data.year or datetime.now().year
    
    previous = db.query(Fair).filter(Fair.name.ilike(fair_data.name or name), Fair.year < year).order_by(Fair.year.desc()).first()
    previous_editions = []
    if previous:
        previous_editions = [{"id": previous.id, "year": previous.year, "status": previous.status, "recommendation": previous.recommendation}]
    
    fair = Fair(
        id=fair_id,
        name=fair_data.name or name,
        year=year,
        url=fair_data.fair_url,
        company_website=fair_data.site_url or fair_data.fair_url,
        company_linkedin=fair_data.linkedin_url or None,
        fair_email=fair_data.fair_email or None,
        folder_path=fair_data.folder_path or None,
        network_path=fair_data.network_path or None,
        status=fair_data.status or "in_valutazione",
        dates=fair_data.dates,
        location=fair_data.location,
        venue=fair_data.venue,
        address=fair_data.address,
        sector=fair_data.sector,
        organizer=fair_data.organizer,
        frequency=fair_data.frequency,
        edition=fair_data.edition,
        expected_visitors=fair_data.expected_visitors,
        exhibitors_count=fair_data.exhibitors_count,
        stand_cost=fair_data.stand_cost,
        exhibitor_countries=fair_data.exhibitor_countries,
        visitor_profile=fair_data.visitor_profile,
        target_segments=fair_data.target_segments,
        product_categories=fair_data.product_categories,
        key_features=fair_data.key_features,
        description=fair_data.description,
        instagram=fair_data.instagram,
        facebook=fair_data.facebook,
        tiktok=fair_data.tiktok,
        contacts=fair_data.contacts,
        previous_editions=previous_editions,
        sources=[{"type": "manual", "url": fair_data.fair_url, "date": datetime.now().isoformat()}]
    )
    db.add(fair)
    db.commit()
    db.refresh(fair)
    return {"id": fair.id, "name": fair.name, "url": fair.url, "status": "created"}


@app.get("/api/fairs", response_model=list[dict])
def list_fairs(skip: int = 0, limit: int = 100, show_archived: bool = False, db: Session = Depends(get_db)):
    query = db.query(Fair)
    if not show_archived:
        query = query.filter(Fair.archived != "yes")
    fairs = query.order_by(Fair.id.desc()).offset(skip).limit(limit).all()
    return [{"id": f.id, "name": f.name or "N/A", "year": f.year, "url": f.url or "", "description": f.description or "", "folder_path": f.folder_path or "", "site_url": f.company_website or "", "linkedin_url": f.company_linkedin or "", "fair_email": f.fair_email or "", "gallery": f.gallery or [], "attachments": f.attachments or [], "contacts": f.contacts or {}, "stand_cost": f.stand_cost or 0, "status": f.status or "in_valutazione", "scraped_data": f.scraped_data, "recommendation": f.recommendation or "", "instagram": f.instagram or "", "facebook": f.facebook or "", "tiktok": f.tiktok or "", "archived": f.archived or "no"} for f in fairs]


@app.get("/api/fairs/{fair_id}", response_model=dict)
def get_fair(fair_id: str, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    return {"id": fair.id, "name": fair.name, "year": fair.year, "url": fair.url, "description": fair.description or "", "folder_path": fair.folder_path or "", "site_url": fair.company_website, "dates": fair.dates, "location": fair.location, "target_segments": fair.target_segments, "expected_visitors": fair.expected_visitors, "exhibitors_count": fair.exhibitors_count, "sources": fair.sources, "web_sources": fair.web_sources or [], "extraction_regions": fair.extraction_regions or [], "linkedin_url": fair.company_linkedin, "fair_email": fair.fair_email or "", "gallery": fair.gallery or [], "attachments": fair.attachments or [], "contacts": fair.contacts or {}, "stand_cost": fair.stand_cost or 0, "status": fair.status or "in_valutazione", "archived": fair.archived or "no", "scraped_data": fair.scraped_data, "historical_data": fair.historical_data, "ROI_assessment": fair.ROI_assessment, "cost_estimate": fair.cost_estimate, "recommendation": fair.recommendation, "rationale": fair.rationale, "report_pdf_path": fair.report_pdf_path, "report_html_path": fair.report_html_path, "venue": fair.venue, "address": fair.address, "sector": fair.sector, "frequency": fair.frequency, "edition": fair.edition, "organizer": fair.organizer, "exhibitor_countries": fair.exhibitor_countries, "visitor_profile": fair.visitor_profile, "product_categories": fair.product_categories, "key_features": fair.key_features, "instagram": fair.instagram or "", "facebook": fair.facebook or "", "tiktok": fair.tiktok or ""}


@app.put("/api/fairs/{fair_id}", response_model=dict)
def update_fair(fair_id: str, fair_data: FairUpdate, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    update_data = fair_data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(fair, key, value)
    db.commit()
    db.refresh(fair)
    return {"id": fair.id, "status": "updated"}


def extract_local_fair_data(title_tag, meta_desc, h1_texts, h2_texts, all_text, emails, phones, future_dates, venue_candidates) -> dict:
    """Algoritmo locale per estrarre dati della fiera senza AI."""
    data = {"name": "", "next_date": "", "organizer": "", "sector": "", "venue": "", "location": "", "email": "", "phone": "", "expected_visitors": 0, "exhibitors_count": 0, "stand_cost": 0, "frequency": "", "summary": ""}
    
    if title_tag:
        data["name"] = title_tag.text.strip() if title_tag else ""
    
    if h1_texts and len(h1_texts) > 0:
        if not data["name"] or data["name"] == title_tag.text.strip() if title_tag else "":
            data["name"] = h1_texts[0]
    
    if future_dates and len(future_dates) > 0:
        data["next_date"] = future_dates[0]
    
    if venue_candidates and len(venue_candidates) > 0:
        data["location"] = venue_candidates[0]
    
    if emails and len(emails) > 0:
        data["email"] = emails[0]
    
    if phones and len(phones) > 0:
        data["phone"] = phones[0]
    
    text_lower = all_text.lower()
    
    sectors = {
        "food": ["food", "cibo", "alimentation", "gastronomia", "agricoltura"],
        "tech": ["tech", "technology", "digital", "ict", "software", "innovation"],
        "build": ["costruzioni", "building", "construction", "real estate", "architettura"],
        "health": ["health", "salute", "medical", "pharma", "farmaceutico"],
        "textile": ["textile", "tessile", "fashion", "abbigliamento", "moda"],
        "logistics": ["logistics", "logistica", "trasporti", "shipping"],
        "tourism": ["turismo", "tourism", "hospitality", "alberghiero"],
        "energy": ["energy", "energia", "green", "renewable", "rinnovabili"],
        "auto": ["auto", "automotive", "motors", "veicoli"],
    }
    for sector, keywords in sectors.items():
        for kw in keywords:
            if kw in text_lower:
                data["sector"] = sector.capitalize()
                break
        if data["sector"]:
            break
    
    freq_map = {"annuale": "annuale", "annual": "annuale", "biennale": "biennale", "biennial": "biennale", "triennale": "triennale", "triennial": "triennale"}
    for freq, kw in freq_map.items():
        if freq in text_lower:
            data["frequency"] = kw
            break
    
    import re
    visitors_match = re.search(r"(\d{1,3}(?:[\.\s]\d{3})*)\s*(?:visitatori|visitors|affluenza|visitatori_previsti)", all_text, re.IGNORECASE)
    if visitors_match:
        data["expected_visitors"] = int(visitors_match.group(1).replace(".", "").replace(" ", ""))
    
    exhibitors_match = re.search(r"(\d{1,3}(?:[\.\s]\d{3})*)\s*(?:espositori|exhibitors|stand)", all_text, re.IGNORECASE)
    if exhibitors_match:
        data["exhibitors_count"] = int(exhibitors_match.group(1).replace(".", "").replace(" ", ""))
    
    cost_match = re.search(r"(\d{1,3}(?:[\.\s]\d{3})*)\s*(?:€|euro)", all_text, re.IGNORECASE)
    if cost_match:
        data["stand_cost"] = int(cost_match.group(1).replace(".", "").replace(" ", ""))
    
    org_patterns = [
        r"organizzato\s+da\s+([A-Za-z][A-Za-z\s]{2,40})",
        r"organizer[:\s]+([A-Za-z][A-Za-z\s]{2,40})",
        r"by\s+([A-Za-z][A-Za-z\s]{2,40})",
    ]
    for pat in org_patterns:
        match = re.search(pat, all_text, re.IGNORECASE)
        if match:
            data["organizer"] = match.group(1).strip()[:50]
            break
    
    venue_patterns = [
        r"(?:fiera|centro\s+fieristico|quartiere)\s+([A-Za-z][A-Za-z\s]{2,40})",
        r"venue[:\s]+([A-Za-z][A-Za-z\s]{2,40})",
    ]
    for pat in venue_patterns:
        match = re.search(pat, all_text, re.IGNORECASE)
        if match:
            data["venue"] = match.group(1).strip()[:50]
            break
    
    if meta_desc:
        desc = meta_desc.get("content", "")[:500]
        if desc:
            data["summary"] = f"Fiera: {data['name'] or 'N/A'}. {desc[:300]} Location: {data['location'] or 'N/A'}. Settore: {data['sector'] or 'N/A'}."
    
    if not data["summary"] or len(data["summary"]) < 50:
        summary_parts = []
        if data["name"]:
            summary_parts.append(f" {data['name']}")
        if data["next_date"]:
            summary_parts.append(f" Data: {data['next_date']}")
        if data["location"]:
            summary_parts.append(f" Luogo: {data['location']}")
        if data["sector"]:
            summary_parts.append(f" Settore: {data['sector']}")
        if data["expected_visitors"]:
            summary_parts.append(f" Visitatori attesi: {data['expected_visitors']}")
        if data["exhibitors_count"]:
            summary_parts.append(f" Espositori: {data['exhibitors_count']}")
        data["summary"] = ". ".join(summary_parts)
        if data["location"] or data["sector"]:
            data["summary"] += ". Per valutare la partecipazione analizzare costi e benefici."
    
    return data


@app.delete("/api/fairs/{fair_id}", response_model=dict)
def delete_fair(fair_id: str, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    db.query(FairAnalysis).filter(FairAnalysis.fair_id == fair_id).delete()
    db.query(OfferComponent).filter(OfferComponent.fair_id == fair_id).delete()
    db.delete(fair)
    db.commit()
    return {"status": "deleted", "id": fair_id}


@app.get("/api/fairs/{fair_id}/analyses", response_model=list[dict])
def list_fair_analyses(fair_id: str, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    analyses = db.query(FairAnalysis).filter(FairAnalysis.fair_id == fair_id).order_by(FairAnalysis.created_at.desc()).all()
    return [{"id": a.id, "name": a.name, "parameters": a.parameters, "result": a.result, "summary": a.summary, "created_at": a.created_at} for a in analyses]


class AnalysisCreate(BaseModel):
    name: str | None = None
    parameters: dict | None = None
    result: dict | None = None
    summary: str | None = None


@app.post("/api/fairs/{fair_id}/analyses", response_model=dict)
def create_fair_analysis(fair_id: str, data: AnalysisCreate, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    analysis = FairAnalysis(
        fair_id=fair_id,
        name=data.name,
        parameters=data.parameters,
        result=data.result,
        summary=data.summary,
        created_at=datetime.now().isoformat()
    )
    db.add(analysis)
    db.commit()
    db.refresh(analysis)
    return {"id": analysis.id, "status": "created"}


@app.delete("/api/fairs/{fair_id}/analyses/{analysis_id}", response_model=dict)
def delete_fair_analysis(fair_id: str, analysis_id: int, db: Session = Depends(get_db)):
    analysis = db.query(FairAnalysis).filter(FairAnalysis.id == analysis_id, FairAnalysis.fair_id == fair_id).first()
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")
    db.delete(analysis)
    db.commit()
    return {"status": "deleted", "id": analysis_id}


@app.get("/api/fairs/{fair_id}/components", response_model=list[dict])
def list_fair_components(fair_id: str, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    components = db.query(OfferComponent).filter(OfferComponent.fair_id == fair_id).all()
    return [{"id": c.id, "name": c.name, "category": c.category, "description": c.description, "quantity": c.quantity, "unit_price": c.unit_price, "total_price": c.total_price, "notes": c.notes} for c in components]


class ComponentCreate(BaseModel):
    name: str
    category: str | None = None
    description: str | None = None
    quantity: int | None = None
    unit_price: float | None = None
    notes: str | None = None


@app.post("/api/fairs/{fair_id}/components", response_model=dict)
def create_fair_component(fair_id: str, data: ComponentCreate, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    total_price = (data.quantity or 1) * (data.unit_price or 0)
    component = OfferComponent(
        fair_id=fair_id,
        name=data.name,
        category=data.category,
        description=data.description,
        quantity=data.quantity,
        unit_price=data.unit_price,
        total_price=total_price,
        notes=data.notes,
        created_at=datetime.now().isoformat()
    )
    db.add(component)
    db.commit()
    db.refresh(component)
    return {"id": component.id, "status": "created"}


@app.put("/api/fairs/{fair_id}/components/{comp_id}", response_model=dict)
def update_fair_component(fair_id: str, comp_id: int, data: ComponentCreate, db: Session = Depends(get_db)):
    component = db.query(OfferComponent).filter(OfferComponent.id == comp_id, OfferComponent.fair_id == fair_id).first()
    if not component:
        raise HTTPException(status_code=404, detail="Component not found")
    component.name = data.name
    component.category = data.category
    component.description = data.description
    component.quantity = data.quantity
    component.unit_price = data.unit_price
    component.total_price = (data.quantity or 1) * (data.unit_price or 0)
    component.notes = data.notes
    db.commit()
    return {"id": component.id, "status": "updated"}


@app.delete("/api/fairs/{fair_id}/components/{comp_id}", response_model=dict)
def delete_fair_component(fair_id: str, comp_id: int, db: Session = Depends(get_db)):
    component = db.query(OfferComponent).filter(OfferComponent.id == comp_id, OfferComponent.fair_id == fair_id).first()
    if not component:
        raise HTTPException(status_code=404, detail="Component not found")
    db.delete(component)
    db.commit()
    return {"status": "deleted", "id": comp_id}


@app.get("/api/fairs/{fair_id}/contacts", response_model=list[dict])
def list_fair_contacts(fair_id: str, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    contacts = fair.contact_list
    return [{"id": c.id, "name": c.name, "email": c.email, "phone": c.phone, "company": c.company, "role": c.role} for c in contacts]


class ContactCreate(BaseModel):
    name: str
    email: str | None = None
    phone: str | None = None
    company: str | None = None
    role: str | None = None
    notes: str | None = None


@app.post("/api/fairs/{fair_id}/contacts", response_model=dict)
def add_fair_contact(fair_id: str, data: ContactCreate, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    contact = Contact(
        name=data.name,
        email=data.email,
        phone=data.phone,
        company=data.company,
        role=data.role,
        notes=data.notes,
        created_at=datetime.now().isoformat()
    )
    fair.contact_list.append(contact)
    db.commit()
    return {"id": contact.id, "status": "added"}


@app.delete("/api/fairs/{fair_id}/contacts/{contact_id}", response_model=dict)
def remove_fair_contact(fair_id: str, contact_id: int, db: Session = Depends(get_db)):
    contact = db.query(Contact).filter(Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if contact in fair.contact_list:
        fair.contact_list.remove(contact)
    db.commit()
    return {"status": "removed", "id": contact_id}


@app.post("/api/fairs/{fair_id}/proposals", response_model=dict)
async def upload_proposal(fair_id: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    
    file_id = str(uuid4())
    file_path = PROPOSALS_DIR / f"{file_id}_{file.filename}"
    with open(file_path, "wb") as f:
        f.write(await file.read())
    
    proposal = CommercialProposal(
        fair_id=fair_id,
        name=file.filename,
        file_path=str(file_path),
        file_name=file.filename,
        received_at=datetime.now().isoformat()
    )
    
    try:
        import fitz
        doc = fitz.open(file_path)
        text = "\n".join([page.get_text() for page in doc])
        doc.close()
        
        import re
        amount_match = re.search(r"(?:€|euro|EUR)\s*(\d{1,3}(?:[.,]\d{3})*)", text, re.IGNORECASE)
        if amount_match:
            proposal.total_amount = float(amount_match.group(1).replace(".", "").replace(",", "."))
        
        stand_match = re.search(r"(\d+)\s*(?:mq|m2|metri quadri)", text, re.IGNORECASE)
        if stand_match:
            proposal.stand_size = int(stand_match.group(1))
        
        location_kws = ["angolo", "centro", "ingresso", "padiglione", "hall"]
        for kw in location_kws:
            if kw in text.lower():
                proposal.stand_location = kw
                break
    except Exception:
        pass
    
    db.add(proposal)
    db.commit()
    db.refresh(proposal)
    return {"id": proposal.id, "status": "uploaded"}


@app.get("/api/fairs/{fair_id}/proposals", response_model=list[dict])
def list_fair_proposals(fair_id: str, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    proposals = db.query(CommercialProposal).filter(CommercialProposal.fair_id == fair_id).all()
    return [{"id": p.id, "name": p.name, "file_name": p.file_name, "total_amount": p.total_amount, "stand_size": p.stand_size, "stand_location": p.stand_location, "status": p.status, "received_at": p.received_at} for p in proposals]


@app.delete("/api/fairs/{fair_id}/proposals/{proposal_id}", response_model=dict)
def delete_proposal(fair_id: str, proposal_id: int, db: Session = Depends(get_db)):
    proposal = db.query(CommercialProposal).filter(CommercialProposal.id == proposal_id, CommercialProposal.fair_id == fair_id).first()
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    db.delete(proposal)
    db.commit()
    return {"status": "deleted", "id": proposal_id}


@app.post("/api/scrape-url")
def scrape_url(url_data: dict, db: Session = Depends(get_db)):
    url = url_data.get("url", "")
    fair_id = url_data.get("fair_id")
    if not url:
        raise HTTPException(status_code=400, detail="URL required")

    from datetime import datetime

    attachments_text = ""
    if fair_id:
        fair = db.query(Fair).filter(Fair.id == fair_id).first()
        if fair and fair.attachments:
            all_text = []
            for att in fair.attachments:
                att_url = att.get("url", "") if isinstance(att, dict) else str(att)
                if not att_url:
                    continue
                full_path = Path("." + att_url) if att_url.startswith("/") else Path(att_url)
                if full_path.exists():
                    ext = full_path.suffix.lower()
                    try:
                        if ext == ".pdf":
                            import fitz
                            doc = fitz.open(full_path)
                            text = "\n".join([page.get_text() for page in doc])
                            doc.close()
                            all_text.append(f"\n=== {full_path.name} ===\n{text[:3000]}")
                        elif ext in [".txt", ".md"]:
                            text = full_path.read_text(encoding="utf-8", errors="ignore")
                            all_text.append(f"\n=== {full_path.name} ===\n{text[:3000]}")
                    except Exception:
                        pass
            attachments_text = "\n\n".join(all_text)

    result = {
        "url": url,
        "fair_id": fair_id,
        "title": "",
        "description": "",
        "text_content": "",
        "summary": "",
        "error": "",
        "ai_data": None,
        "logs": [],
        "attachments_text": attachments_text[:2000] if attachments_text else ""
    }

    def log(step: str, message: str, status: str = "ok"):
        result["logs"].append({
            "ts": datetime.now().strftime("%H:%M:%S"),
            "step": step,
            "message": message,
            "status": status
        })

    import re

    try:
        log("START", f"Inizio raccolta dati da {url}")

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept-Language": "en-US,en;q=0.9,it-IT;q=0.8,it;q=0.7",
        }
        resp = requests.get(url, timeout=20, verify=False, headers=headers, allow_redirects=True)
        log("HTTP", f"Status: {resp.status_code}", "ok" if resp.status_code == 200 else "warning")

        html_content = resp.text
        if not html_content:
            result["error"] = "No content"
            log("ERROR", "Nessun contenuto ricevuto", "error")
            return result

        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html_content, "html.parser")

        title = soup.find("title")
        result["title"] = title.text.strip() if title else ""
        if result["title"]:
            log("TITLE", f"Titolo: {result['title'][:60]}")

        meta_desc = soup.find("meta", attrs={"name": "description"})
        if meta_desc:
            result["description"] = meta_desc.get("content", "")[:500]
            log("META", "Meta description estratta")

        for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
            tag.decompose()

        h1_texts = [h.get_text(strip=True) for h in soup.find_all("h1")]
        h2_texts = [h.get_text(strip=True) for h in soup.find_all("h2")]
        h3_texts = [h.get_text(strip=True) for h in soup.find_all("h3")]
        
        structured_text = "\n".join([
            "TITOLI H1: " + " | ".join(h1_texts[:5]),
            "TITOLI H2: " + " | ".join(h2_texts[:10]),
            "TITOLI H3: " + " | ".join(h3_texts[:15]),
        ])

        contact_patterns = [
            r"[\w\.-]+@[\w\.-]+\.\w+",
            r"\+?\d{1,3}[-.\s]?\(?\d{1,4}\)?[-.\s]?\d{1,4}[-.\s]?\d{1,9}",
            r"tel[:\s]*(\+?[\d\s\-\(\)]+)",
            r"phone[:\s]*(\+?[\d\s\-\(\)]+)",
        ]
        all_text = soup.get_text(separator=" ", strip=True)
        emails = list(set(re.findall(r"[\w\.-]+@[\w\.-]+\.\w+", all_text)))
        phones = list(set(re.findall(r"\+?\d{1,3}[-.\s]?\(?\d{1,4}\)?[-.\s]?\d{1,4}[-.\s]?\d{1,9}", all_text)))
        
        contact_info = []
        if emails:
            contact_info.append("Email: " + ", ".join(emails[:5]))
        if phones:
            contact_info.append("Tel: " + ", ".join(phones[:3]))

        date_patterns = [
            r"(\d{1,2}\s+(?:gen|feb|mar|apr|mag|giu|lug|ago|set|ott|nov|dic)\w*\s+\d{4})",
            r"(\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4})",
            r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
            r"(\d{4})",
        ]
        all_dates = []
        current_year = datetime.now().year
        for pat in date_patterns:
            matches = re.findall(pat, all_text, re.IGNORECASE)
            for m in matches:
                year_match = re.search(r"\d{4}", m)
                if year_match:
                    year = int(year_match.group())
                    if year >= current_year:
                        all_dates.append(m)
        
        future_dates = list(set(all_dates))[:10]

        venue_cities = ["milano", "roma", "torino", "bologna", "firenze", "napoli", "genova", "verona", "padova", "vicenza", "bari", "catania", "rimini", "modena", "pesaro"]
        venue_patterns = [
            r"(?:fiera|centro\s+fieristico|centro\s+congressi|quartiere\s+fieristico)\s+([^,\.]+)",
            r"(?:presso|a|in)\s+([^,\.]+?(?:fiera|centro|palazzo|hub))",
        ]
        venues_found = []
        for pat in venue_patterns:
            matches = re.findall(pat, all_text, re.IGNORECASE)
            venues_found.extend([m.strip()[:50] for m in matches])
        
        venue_candidates = []
        text_lower = all_text.lower()
        for city in venue_cities:
            if city in text_lower:
                venue_candidates.append(city.capitalize())
        venue_candidates.extend(list(set(venues_found))[:3])

        text_content = " ".join(all_text.split())
        result["text_content"] = text_content[:15000]
        log("HTML", f"Testo estratto: {len(text_content)} char")

        structured_data = f"""
=== DATI ESTRATTI DALLA PAGINA ===

TITOLO: {result['title']}

DESCRIZIONE META: {result['description']}

TITOLI PAGINA:
{structured_text}

DATE FUTURE TROVATE: {', '.join(future_dates) if future_dates else 'N/A'}

CITTA'/LOCATION: {', '.join(venue_candidates) if venue_candidates else 'N/A'}

CONTATTI: {' | '.join(contact_info) if contact_info else 'N/A'}

CONTENUTO PAGINA (primi 3000 char):
{text_content[:3000]}

ALLEGATI (estratti da PDF/documenti):
{result['attachments_text'][:3000] if result['attachments_text'] else 'Nessun allegato'}
"""

        settings = db.query(Settings).first()
        use_ai = settings and settings.ollama_model
        ollama_available = False
        
        if use_ai:
            try:
                test_resp = requests.get(f"{settings.ollama_url}/api/tags", timeout=2)
                ollama_available = test_resp.status_code == 200
            except Exception:
                pass
        
        if use_ai and ollama_available:
            log("OLLAMA", f"Analisi con Ollama: {settings.ollama_model}")
            try:
                client = OllamaClient(settings.ollama_url)
                model_name = str(settings.ollama_model)

                prompt = f"""Sei un esperto analyst di fiere commerciali. Analizza questa pagina web di una fiera e i suoi allegati (PDF/documenti), estrai tutte le informazioni rilevanti e combinale per ottenere dati completi e accurati.

{structured_data}

Istruzioni:
1. Combina e verifica i dati trovati nella pagina web e negli allegati
2. In caso di discrepanze, usa i dati più recenti/aggiornati (gli allegati hanno priorità per dati ufficiali)
3. Estrai il NOME UFFICIALE della fiera
4. Identifica la PROSSIMA DATA della fiera (la più vicina nel futuro)
5. Trova l'ORGANIZZATORE della fiera  
6. Identifica il SETTORE merceologico principale
7. Trova il LUOGO/VENUE dove si tiene
8. Estrai tutti i DATI DI CONTATTO (email, telefono, indirizzo)
9. Estrai il NUMERO VISITATORI previsti (se indicato)
10. Estrai il NUMERO ESPOSITORI (se indicato)
11. Estrai il COSTO dello stand (se indicato)
12. Estrai la FREQUENZA della fiera (annuale, biennale, triennale)
13. Scrivi un RIASSUNTO di almeno 500 parole sulla fiera basandoti sul contenuto combinato di pagina web e allegati

Rispondi in JSON:
{{
  "name": "nome fiera",
  "next_date": "data prossima edizione",
  "organizer": "nome organizzatore",
  "sector": "settore merceologico",
  "venue": "nome venue",
  "location": "città",
  "email": "email contatto",
  "phone": "telefono",
  "expected_visitors": numero,
  "exhibitors_count": numero,
  "stand_cost": numero,
  "frequency": "annuale|biennale|triennale",
  "summary": "riassunto di almeno 500 parole sulla fiera"
}}

Rispondi SOLO con JSON valido, niente altro testo."""

                ai_resp = client.chat(model_name, prompt)
                if ai_resp:
                    match = re.search(r"\{[\s\S]*\}", ai_resp)
                    if match:
                        import json as json_lib
                        ai_data = json_lib.loads(match.group())
                        result["ai_data"] = ai_data
                        
                        for k, v in ai_data.items():
                            if v:
                                log("AI", f"{k}: {str(v)[:50]}")
                        
                        if ai_data.get("summary"):
                            result["summary"] = ai_data["summary"]
                            log("SUMMARY", f"Riassunto generato: {len(ai_data['summary'])} parole")
                    else:
                        log("AI", "Nessun JSON - uso algoritmo locale", "warning")
                        result["ai_data"] = extract_local_fair_data(title, meta_desc, h1_texts, h2_texts, all_text, emails, phones, future_dates, venue_candidates)
                else:
                    log("AI", "Nessuna risposta - uso algoritmo locale", "warning")
                    result["ai_data"] = extract_local_fair_data(title, meta_desc, h1_texts, h2_texts, all_text, emails, phones, future_dates, venue_candidates)

            except Exception as e:
                log("AI_ERROR", str(e)[:60], "warning")
                log("LOCAL", "Fallback algoritmo locale")
                result["ai_data"] = extract_local_fair_data(title, meta_desc, h1_texts, h2_texts, all_text, emails, phones, future_dates, venue_candidates)
        else:
            log("LOCAL", "Ollama non disponibile - uso algoritmo locale")
            result["ai_data"] = extract_local_fair_data(title, meta_desc, h1_texts, h2_texts, all_text, emails, phones, future_dates, venue_candidates)

        if not result["title"]:
            result["title"] = result.get("ai_data", {}).get("name", "")
        if not result["description"] and text_content:
            result["description"] = text_content[:400]

    except Exception as e:
        result["error"] = str(e)
        log("ERROR", str(e)[:60], "error")

    log("END", "Raccolta completata")
    return result


@app.post("/api/fairs/{fair_id}/analyze")
def analyze_fair(fair_id: str, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")

    broadcast_notification(f"Avvio analisi per {fair.name}...", "info", fair_id)
    background_tasks.add_task(run_analyze_fair, fair_id)
    
    return {"status": "started", "fair_id": fair_id, "message": "Analisi avviata in background"}


async def run_analyze_fair(fair_id: str):
    from .db import SessionLocal
    from .models import Fair
    
    db = SessionLocal()
    try:
        fair = db.query(Fair).filter(Fair.id == fair_id).first()
        if not fair:
            broadcast_notification("Fiera non trovata", "error", fair_id)
            return
        
        attachments = fair.attachments or []
        attachments_text = ""
        
        if attachments:
            all_text = []
            for att in attachments:
                att_url = att.get("url", "") if isinstance(att, dict) else str(att)
                if not att_url:
                    continue
                full_path = Path("." + att_url) if att_url.startswith("/") else Path(att_url)
                if full_path.exists():
                    ext = full_path.suffix.lower()
                    try:
                        if ext == ".pdf":
                            import fitz
                            doc = fitz.open(full_path)
                            text = "\n".join([page.get_text() for page in doc])
                            doc.close()
                            all_text.append(f"\n=== {full_path.name} ===\n{text[:2000]}")
                            broadcast_notification(f"Estratto PDF {full_path.name}", "info", fair_id)
                        elif ext in [".txt", ".md"]:
                            text = full_path.read_text(encoding="utf-8", errors="ignore")
                            all_text.append(f"\n=== {full_path.name} ===\n{text[:2000]}")
                    except Exception as e:
                        broadcast_notification(f"Errore: {str(e)}", "warning", fair_id)
            attachments_text = "\n\n".join(all_text)[:5000]
        
        fair.scraped_data = {
            "analyzed_at": datetime.now().isoformat(),
            "attachments_text": attachments_text[:500] if attachments_text else None
        }
        db.commit()
        
        broadcast_notification(f"Analisi completata per {fair.name}", "success", fair_id)
        
    except Exception as e:
        broadcast_notification(f"Errore analisi: {str(e)}", "error", fair_id)
    finally:
        db.close()


def analyze_fair_task(fair_id: str):
    from .db import SessionLocal
    from .models import Fair, Settings
    from .services.ollama import OllamaClient
    import requests
    from bs4 import BeautifulSoup
    from datetime import datetime

    db = SessionLocal()
    try:
        fair = db.query(Fair).filter(Fair.id == fair_id).first()
        if not fair:
            return

        settings = db.query(Settings).first()
        if not settings:
            settings = Settings()

        scraped_data = {}
        url_to_check = fair.url or fair.company_website or ""
        if url_to_check:
            try:
                resp = requests.get(url_to_check, timeout=15, verify=False, headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                    'Accept-Language': 'en-US,en;q=0.9,it-IT;q=0.8,it;q=0.7'
                })
                if resp.status_code == 200:
                    soup = BeautifulSoup(resp.text, "html.parser")
                    scraped_data['title'] = soup.find('title').text.strip() if soup.find('title') else ''
                    scraped_data['source_url'] = url_to_check
                    scraped_data['scraped_at'] = datetime.now().isoformat()
            except Exception as e:
                scraped_data['error'] = str(e)

        scraped_data['final_data'] = {}
        scraped_data['extracted_at'] = datetime.now().isoformat()

        fair.scraped_data = scraped_data
        if not fair.recommendation:
            fair.status = "in_valutazione"
        db.commit()

    except Exception:
        pass
    finally:
        db.close()


@app.post("/api/fairs/{fair_id}/evaluate")
def evaluate_fair(fair_id: str, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")

    attachments = fair.attachments
    has_attachments = False
    if attachments:
        if isinstance(attachments, list):
            has_attachments = len(attachments) > 0
        elif isinstance(attachments, str) and attachments.strip():
            has_attachments = True

    if not has_attachments:
        raise HTTPException(status_code=400, detail="Caricare preventivi e allegati prima della valutazione")

    settings = get_settings_db(db)
    ollama_available = False
    try:
        resp = requests.get(f"{settings.ollama_url}/api/tags", timeout=3)
        ollama_available = resp.status_code == 200
    except Exception:
        pass

    recommendation = None
    rationale = None
    roi = None

    if ollama_available and settings.ollama_model:
        try:
            client = OllamaClient(settings.ollama_url)
            model_name = str(settings.ollama_model) if settings.ollama_model else "phi:latest"

            prompt = f"""You are an expert in trade shows and events. Evaluate this trade fair based on the company strategy AND the attached quotes/documents.

STRATEGY: {settings.strategy_prompt}

FAIR DATA:
- Name: {fair.name}
- URL: {fair.url}
- Description: {fair.description}
- Location: {fair.location}
- Dates: {fair.dates}

Reply ONLY in JSON format:
{{"recommendation": "...", "rationale": "...", "roi_assessment": "..."}}"""

            ai_response = client.chat(model_name, prompt)
            if ai_response:
                import re
                import json
                match = re.search(r'\{[^}]+\}', ai_response, re.DOTALL)
                if match:
                    result = json.loads(match.group())
                    recommendation = result.get('recommendation', '')
                    rationale = result.get('rationale', '')
                    roi = result.get('roi_assessment', '')
        except Exception:
            pass

    if recommendation:
        fair.recommendation = recommendation
    if rationale:
        fair.rationale = rationale
    if roi:
        fair.ROI_assessment = {"assessment": roi, "evaluated_at": datetime.now().isoformat()}

    db.commit()
    return {"status": "evaluated", "ollama_available": ollama_available, "recommendation": recommendation, "rationale": rationale}


class ReportFormat(BaseModel):
    format: str = "html"


@app.post("/api/fairs/{fair_id}/report")
def generate_report(fair_id: str, fmt: ReportFormat = ReportFormat(format="html"), db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    settings = get_settings_db(db)
    from jinja2 import Environment, FileSystemLoader
    env = Environment(loader=FileSystemLoader("./src/fair_evaluator/templates"))
    template = env.get_template("report.html")
    html_content = template.render(fair=fair, strategy_prompt=settings.strategy_prompt or "")
    html_path = REPORTS_DIR / f"report_{fair_id}.html"
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    fair.report_html_path = str(html_path)
    pdf_path = None
    if fmt.format.lower() == 'pdf':
        try:
            from weasyprint import HTML
            pdf_path = str(REPORTS_DIR / f"report_{fair_id}.pdf")
            HTML(string=html_content).write_pdf(pdf_path)
            fair.report_pdf_path = pdf_path
        except Exception:
            pass
    db.commit()
    return {"html": str(html_path), "pdf": pdf_path}


@app.get("/api/fairs/{fair_id}/report/download")
def download_report(fair_id: str, format: str = "html", db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    if format.lower() == "pdf" and fair.report_pdf_path:
        from pathlib import Path
        if Path(fair.report_pdf_path).exists():
            return FileResponse(fair.report_pdf_path, filename=f"report_{fair_id}.pdf")
    if fair.report_html_path:
        from pathlib import Path
        if Path(fair.report_html_path).exists():
            return FileResponse(fair.report_html_path, filename=f"report_{fair_id}.html")
    raise HTTPException(status_code=404, detail="Report not found")


class WebSourceInput(BaseModel):
    url: str
    source_type: str
    label: str | None = None


class WebSourceUpdate(BaseModel):
    sources: list


SCREENSHOTS_DIR = Path("./data/screenshots")


@app.post("/api/fairs/{fair_id}/web-sources")
def add_web_source(fair_id: str, web_source: WebSourceInput, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    
    sources = fair.web_sources or []
    source_entry = {
        "url": web_source.url,
        "source_type": web_source.source_type,
        "label": web_source.label or web_source.source_type,
        "screenshot": None
    }
    sources.append(source_entry)
    fair.web_sources = sources
    db.commit()
    
    return {"sources": sources}


@app.put("/api/fairs/{fair_id}/web-sources")
def update_web_sources(fair_id: str, update: WebSourceUpdate, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    
    fair.web_sources = update.sources
    db.commit()
    
    return {"fair_id": fair_id, "sources": fair.web_sources or []}


@app.post("/api/fairs/{fair_id}/web-sources/screenshot-fallback")
def capture_screenshots_fallback(fair_id: str, url_idx: int, db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    
    sources = fair.web_sources or []
    if url_idx < 0 or url_idx >= len(sources):
        raise HTTPException(status_code=400, detail="Invalid URL index")
    
    url = sources[url_idx].get("url", "")
    if not url:
        raise HTTPException(status_code=400, detail="URL not found")
    
    return {"status": "skipped", "message": "Screenshot disabled - install playwright browsers manually"}


@app.post("/api/fairs/{fair_id}/web-sources/screenshot")
def capture_screenshots(fair_id: str, url_idx: int, db: Session = Depends(get_db)):
    SCREENSHOTS_DIR.mkdir(parents=True, exist_ok=True)
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    
    sources = fair.web_sources or []
    if url_idx < 0 or url_idx >= len(sources):
        raise HTTPException(status_code=400, detail="Invalid URL index")
    
    url = sources[url_idx].get("url", "")
    if not url:
        raise HTTPException(status_code=400, detail="URL not found")
    
    screenshot_path = SCREENSHOTS_DIR / f"{fair_id}_{url_idx}.png"
    
    try:
        import asyncio
        from playwright.async_api import async_playwright
        
        async def take_screenshot():
            async with async_playwright() as p:
                browser = await p.chromium.launch()
                page = await browser.new_page()
                await page.goto(url, timeout=15000)
                await page.screenshot(path=str(screenshot_path), full_page=True)
                await browser.close()
        
        asyncio.run(take_screenshot())
        sources[url_idx]["screenshot"] = str(screenshot_path)
        fair.web_sources = sources
        db.commit()
    except Exception as e:
        return {"error": str(e), "screenshot": None, "status": "error"}
    
    return {"url_idx": url_idx, "screenshot": str(screenshot_path)}


@app.post("/api/fairs/{fair_id}/web-sources/{url_idx}/screenshot-upload")
async def upload_screenshot(fair_id: str, url_idx: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    fair = db.query(Fair).filter(Fair.id == fair_id).first()
    if not fair:
        raise HTTPException(status_code=404, detail="Fair not found")
    
    sources = fair.web_sources or []
    if url_idx < 0 or url_idx >= len(sources):
        raise HTTPException(status_code=400, detail="Invalid URL index")
    
    SCREENSHOTS_DIR.mkdir(parents=True, exist_ok=True)
    screenshot_path = SCREENSHOTS_DIR / f"{fair_id}_{url_idx}_{file.filename}"
    
    contents = await file.read()
    with open(screenshot_path, "wb") as f:
        f.write(contents)
    
    sources[url_idx]["screenshot"] = str(screenshot_path)
    fair.web_sources = sources
    db.commit()
    
    return {"url_idx": url_idx, "screenshot": str(screenshot_path)}


@app.post("/api/import-excel-upload")
async def import_excel_upload(file: UploadFile = File(...), db: Session = Depends(get_db)):
    import openpyxl
    import io
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(h).lower().strip() if h else "" for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        col_idx = {h: i for i, h in enumerate(headers) if h}

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[col_idx.get("nome", 0)]).strip() if col_idx.get("nome") is not None and len(row) > col_idx.get("nome", 0) and row[col_idx.get("nome", 0)] else ""
            url = str(row[col_idx.get("link evento", 0)]).strip() if col_idx.get("link evento") is not None and len(row) > col_idx.get("link evento", 0) and row[col_idx.get("link evento", 0)] else ""

            if not name and not url:
                continue

            if not name and url:
                if '/' in url:
                    name = url.split('/')[-1]
                    if '.' in name:
                        name = name.split('.')[0]

            inizio = str(row[col_idx.get("inizio", 0)]).strip() if col_idx.get("inizio") is not None and len(row) > col_idx.get("inizio", 0) and row[col_idx.get("inizio", 0)] else ""
            dates = [inizio] if inizio else None

            fair = Fair(
                id=str(uuid4()),
                name=name or "Fiera",
                url=url or None,
                location=str(row[col_idx.get("città", 0)]).strip() if col_idx.get("città") is not None and len(row) > col_idx.get("città", 0) and row[col_idx.get("città", 0)] else None,
                venue=str(row[col_idx.get("location", 0)]).strip() if col_idx.get("location") is not None and len(row) > col_idx.get("location", 0) and row[col_idx.get("location", 0)] else None,
                dates=dates,
                folder_path=str(row[col_idx.get("cartella", 0)]).strip() if col_idx.get("cartella") is not None and len(row) > col_idx.get("cartella", 0) and row[col_idx.get("cartella", 0)] else None,
                stand_cost=int(row[col_idx.get("costo stand", 0)]) if col_idx.get("costo stand") is not None and len(row) > col_idx.get("costo stand", 0) and row[col_idx.get("costo stand", 0)] and str(row[col_idx.get("costo stand", 0)]).isdigit() else 0,
                status=str(row[col_idx.get("stato", 0)]).strip() if col_idx.get("stato") is not None and len(row) > col_idx.get("stato", 0) and row[col_idx.get("stato", 0)] else "in_valutazione",
                sources=[{"type": "excel_import", "url": url, "date": datetime.now().isoformat()}]
            )
            db.add(fair)
            count += 1
        db.commit()
        wb.close()
        return {"status": "success", "count": count}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/api/export-excel")
def export_excel(db: Session = Depends(get_db)):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fiere"
    ws.append(["Nome", "Città", "Location", "Stato", "Anno", "Inizio", "Durata", "Cartella", "link evento", "Contatto", "Contatto mail", "Costo stand"])
    fairs = db.query(Fair).all()
    for fair in fairs:
        inizio = fair.dates[0] if fair.dates and len(fair.dates) > 0 else ""
        contatti = fair.contacts or {}
        ws.append([
            fair.name or "",
            fair.location or "",
            fair.venue or "",
            fair.status or "in_valutazione",
            inizio[:4] if inizio else "",
            inizio,
            "",
            fair.folder_path or "",
            fair.url or "",
            contatti.get('name', ''),
            contatti.get('email', ''),
            fair.stand_cost or 0
        ])
    output_file = REPORTS_DIR / f"Fiere_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output_file)
    return FileResponse(str(output_file), filename="Fiere.xlsx")


@app.post("/api/db/clear")
def clear_database(db: Session = Depends(get_db)):
    try:
        db.query(Fair).delete()
        db.commit()
        return {"status": "success", "message": "Database cleared"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/api/db/stats")
def get_database_stats(db: Session = Depends(get_db)):
    total = db.query(Fair).count()
    in_valutazione = db.query(Fair).filter(Fair.status == "in_valutazione").count()
    approvata = db.query(Fair).filter(Fair.status == "approvata").count()
    in_gestione = db.query(Fair).filter(Fair.status == "in_gestione").count()
    conclusa = db.query(Fair).filter(Fair.status == "conclusa").count()
    rifiutata = db.query(Fair).filter(Fair.status == "rifiutata").count()

    fairs = db.query(Fair).all()
    total_contacts = {"cold": 0, "warm": 0, "hot": 0}
    total_stand_cost = 0

    for fair in fairs:
        if fair.contacts:
            total_contacts["cold"] += fair.contacts.get("cold", 0)
            total_contacts["warm"] += fair.contacts.get("warm", 0)
            total_contacts["hot"] += fair.contacts.get("hot", 0)
        if fair.stand_cost:
            total_stand_cost += fair.stand_cost

    return {
        "total": total,
        "in_valutazione": in_valutazione,
        "approvata": approvata,
        "in_gestione": in_gestione,
        "conclusa": conclusa,
        "rifiutata": rifiutata,
        "contacts": total_contacts,
        "total_stand_cost": total_stand_cost
    }


@app.post("/api/strategy/load")
async def load_strategy_prompt(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    prompt_content = content.decode("utf-8")

    settings = get_settings_db(db)
    settings.strategy_prompt = prompt_content
    db.commit()

    prompt_file = STRATEGY_DIR / "generated_prompt.txt"
    with open(prompt_file, "w", encoding="utf-8") as f:
        f.write(prompt_content)

    return {"status": "success", "prompt": prompt_content}


@app.post("/api/strategy/analyze")
def analyze_strategy_manual(db: Session = Depends(get_db)):
    import glob
    pdf_files = glob.glob(str(STRATEGY_DIR / "*.pdf"))
    if not pdf_files:
        return {"status": "error", "message": "Nessun file PDF trovato"}

    settings = get_settings_db(db)
    client = OllamaClient(str(settings.ollama_url) if settings.ollama_url else "http://localhost:11434")

    prompt = f"""Il file PDF della strategia marketing si trova in: {pdf_files[0]};

Genera un prompt dettagliato in italiano per valutare le fiere."""

    try:
        response = client.chat(str(settings.ollama_model) if settings.ollama_model else "phi:latest", prompt)
        if response:
            settings.strategy_prompt = response
            db.commit()
            prompt_file = STRATEGY_DIR / "generated_prompt.txt"
            with open(prompt_file, "w", encoding="utf-8") as f:
                f.write(response)
            return {"status": "success", "prompt": response}
    except Exception as e:
        return {"status": "error", "message": str(e)}

    return {"status": "error", "message": "Impossibile generare il prompt"}


TEMPLATE_DIR = Path(__file__).resolve().parent / "templates"


def render_template(template_name: str, **kwargs) -> str:
    template_path = TEMPLATE_DIR / template_name
    if template_path.exists():
        with open(template_path, "r", encoding="utf-8") as f:
            return f.read()
    return f"<h1>Template {template_name} not found</h1>"


@app.get("/", response_class=HTMLResponse)
def home_page(db: Session = Depends(get_db)):
    from jinja2 import Environment, FileSystemLoader
    env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)))
    template = env.get_template("page_home.html")
    stats = get_database_stats(db)
    ollama = check_ollama_status(db)
    settings = get_settings_endpoint(db)
    return HTMLResponse(template.render(stats=stats, ollama=ollama, settings=settings, nav_active="home"))


@app.get("/fairs", response_class=HTMLResponse)
def fairs_list_page(db: Session = Depends(get_db)):
    from jinja2 import Environment, FileSystemLoader
    env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)))
    template = env.get_template("page_fairs.html")
    fairs = db.query(Fair).filter(Fair.archived != "yes").order_by(Fair.id.desc()).limit(100).all()
    locations = sorted(set(f.location for f in fairs if f.location))
    return HTMLResponse(template.render(fairs=fairs, locations=locations, nav_active="fairs"))


@app.get("/fairs/new", response_class=HTMLResponse)
def new_fair_page():
    from jinja2 import Environment, FileSystemLoader
    env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)))
    template = env.get_template("page_fair_new.html")
    return HTMLResponse(template.render(nav_active="fairs"))


@app.get("/fairs/{fair_id}", response_class=HTMLResponse)
def fair_detail_page(fair_id: str):
    from jinja2 import Environment, FileSystemLoader
    env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)))
    template = env.get_template("page_fair_detail.html")
    fair = get_fair(fair_id, SessionLocal())
    return HTMLResponse(template.render(fair=fair, nav_active="fairs"))


@app.get("/fairs/{fair_id}/visual-editor", response_class=HTMLResponse)
def visual_editor_page(fair_id: str):
    from jinja2 import Environment, FileSystemLoader
    env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)))
    template = env.get_template("page_visual_editor.html")
    return HTMLResponse(template.render(fair_id=fair_id, nav_active="fairs"))


@app.get("/visual-editor", response_class=HTMLResponse)
def visual_editor_standalone(request: Request):
    from jinja2 import Environment, FileSystemLoader
    env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)))
    template = env.get_template("page_visual_editor.html")
    url = request.query_params.get("url", "")
    return HTMLResponse(template.render(url=url, nav_active=""))


@app.get("/settings", response_class=HTMLResponse)
def settings_page(db: Session = Depends(get_db)):
    from jinja2 import Environment, FileSystemLoader
    env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)))
    template = env.get_template("page_settings.html")
    settings = get_settings_endpoint(db)
    ollama = check_ollama_status(db)
    return HTMLResponse(template.render(settings=settings, ollama=ollama, nav_active="settings"))


@app.get("/maintenance", response_class=HTMLResponse)
def maintenance_page(db: Session = Depends(get_db)):
    from jinja2 import Environment, FileSystemLoader
    env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)))
    template = env.get_template("page_maintenance.html")
    stats = get_database_stats(db)
    settings = get_settings_endpoint(db)
    return HTMLResponse(template.render(stats=stats, strategy_prompt=settings.get('strategy_prompt', ''), nav_active="maintenance"))


@app.post("/fair-scan")
def legacy_scan(fair_data: FairCreate):
    return create_fair(fair_data, SessionLocal())